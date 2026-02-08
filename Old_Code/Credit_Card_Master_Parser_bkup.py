import os
import csv
from pathlib import Path
from datetime import datetime
import re
import pdfplumber
import pandas as pd

# Import all parsers
from icici_cc_pdf_parser import extract_icici_transactions
from idfc_cc_pdf_parser import extract_idfc_transactions
from uni_gold_cc_pdf_parser import parse_uni_gold_cc_pdf
from uni_gold_upi_cc_pdf_parser import parse_uni_gold_upi_cc_pdf
from axis_unified_pdf_parser import parse_axis_pdf
from axis_rewards_smart_parser import parse_axis_rewards_smart  # NEW

PROJECT_DIR = os.path.expanduser(
    "~/Library/CloudStorage/OneDrive-Personal/Personal/Finance/projects/Monthly_Fin_Tracker"
)

BASE_DIR = os.path.join(PROJECT_DIR, "CC statements")

OUTPUT_FILE = os.path.join(
    PROJECT_DIR, "Output", "Credit card transactions.xlsx"
)

MAPPING_FILE = os.path.join(
    PROJECT_DIR, "Reference Documents", "Merchant category mapping.csv"
)

PAYMENT_KEYWORDS = [
    "PAYMENT RECEIVED",
    "PAYMENT RECIEVED",
    "SI PAYMENT",
    "SI PAYMENT RECEIVED",
    "BBPS PAYMENT RECEIVED",
    "AUTO-DEBIT",
    "SMS BASED REDEMPTION",
    "PAY BY REWARDS",
    "INFINITY PAYMENT RECEIVED",
]

EXPENSE_TYPE_KEYWORDS = {
    "Shopping": ["AMAZON", "RELIANCE", "MART", "STORE", "FLIPKART", "MYNTRA"],
    "Grocery": ["GROCERY", "SUPERMARKET", "DMART", "BIGBASKET"],
    "Food": ["SWIGGY", "ZOMATO", "DOMINOS", "PIZZA", "RESTAURANT", "CAFE"],
    "Entertainment": ["SPOTIFY", "NETFLIX", "PRIME VIDEO", "HOTSTAR", "BOOKMYSHOW", "PVR", "INOX", "MOVIE"],
    "Travel": ["IXIGO", "IRCTC", "MAKE MY TRIP", "MAKEMYTRIP", "GOIBIBO", "UBER", "OLA", "AIR", "RAIL"],
    "Fuel": ["FUEL", "PETROL", "DIESEL", "INDIAN OIL", "IOCL", "BPCL", "HPCL"],
}


def get_parser(file_path):
    """
    Identify which parser to use based on folder path
    IMPORTANT: Check for Axis Rewards BEFORE general Axis to avoid double parsing
    """
    path = file_path.lower()

    # ICICI
    if "icici" in path:
        return extract_icici_transactions, "ICICI Amazon Pay"

    # IDFC
    if "idfc" in path:
        return extract_idfc_transactions, "IDFC FIRST"

    # Uni Gold UPI (check before regular Uni Gold)
    if "uni gold upi" in path:
        return parse_uni_gold_upi_cc_pdf, "Uni Gold UPI"

    # Uni Gold
    if "uni gold" in path:
        return parse_uni_gold_cc_pdf, "Uni Gold"

    # AXIS - Check Rewards FIRST (most specific)
    if "axis" in path:
        if "rewards" in path:
            return parse_axis_rewards_smart, "Axis Rewards"  # Uses smart parser
        elif "select" in path:
            return parse_axis_pdf, "Axis Select"  # Uses unified parser
        elif "indian oil" in path:
            return parse_axis_pdf, "Axis Indian Oil"  # Uses unified parser
        else:
            return parse_axis_pdf, "Axis Bank"  # Fallback to unified parser

    return None, None


def normalize(records):
    """
    Normalize records - ensure all required fields exist
    """
    valid = []

    for r in records:
        if not isinstance(r, dict):
            continue

        if not r.get("Date"):
            continue

        # Ensure Period exists
        if "Period" not in r or not r["Period"]:
            r["Period"] = "Unknown"

        # Ensure Type exists
        if "Type" not in r or not r["Type"]:
            # Infer from amount
            amount = r.get("Amount", 0)
            if isinstance(amount, str):
                amount = float(amount.replace(",", ""))
            r["Type"] = "Cr" if amount < 0 else "Dr"

        # Enforce sign rules: Dr positive, Cr negative
        amount = r.get("Amount", 0)
        if isinstance(amount, str):
            amount = float(amount.replace(",", ""))
        if str(r.get("Type", "")).upper().startswith("CR") and amount > 0:
            amount = -amount
        if str(r.get("Type", "")).upper().startswith("DR") and amount < 0:
            amount = -amount
        r["Amount"] = amount

        valid.append(r)

    return valid


def ensure_mapping_file():
    """Create or upgrade the mapping file to the new schema."""
    if os.path.exists(MAPPING_FILE):
        return

    starter_rows = [
        {"Keyword Pattern": "AMAZON", "Expense Type": "Shopping", "Merchant Category": "Shopping", "Store Name": "Amazon"},
        {"Keyword Pattern": "RELIANCE", "Expense Type": "Shopping", "Merchant Category": "Shopping", "Store Name": "Reliance"},
        {"Keyword Pattern": "RELIANCE MART", "Expense Type": "Shopping", "Merchant Category": "Shopping", "Store Name": "Reliance Mart"},
        {"Keyword Pattern": "SWIGGY", "Expense Type": "Food", "Merchant Category": "Food Delivery", "Store Name": "Swiggy"},
        {"Keyword Pattern": "ZOMATO", "Expense Type": "Food", "Merchant Category": "Food Delivery", "Store Name": "Zomato"},
        {"Keyword Pattern": "IXIGO", "Expense Type": "Travel", "Merchant Category": "Travel Booking", "Store Name": "Ixigo"},
        {"Keyword Pattern": "SPOTIFY", "Expense Type": "Entertainment", "Merchant Category": "Subscriptions", "Store Name": "Spotify"},
        {"Keyword Pattern": "BBPS PAYMENT RECEIVED", "Expense Type": "Card Payment", "Merchant Category": "CC Bill Payment", "Store Name": "Bank"},
        {"Keyword Pattern": "SMS BASED REDEMPTION", "Expense Type": "Card Payment", "Merchant Category": "CC Bill Payment", "Store Name": "Axis Bank"},
        {"Keyword Pattern": "PAY BY REWARDS", "Expense Type": "Card Payment", "Merchant Category": "CC Bill Payment", "Store Name": "Axis Bank"},
        {"Keyword Pattern": "INFINITY PAYMENT RECEIVED", "Expense Type": "Card Payment", "Merchant Category": "CC Bill Payment", "Store Name": "Axis Bank"},
    ]

    with open(MAPPING_FILE, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(
            f, fieldnames=["Keyword Pattern", "Expense Type", "Merchant Category", "Store Name"]
        )
        writer.writeheader()
        writer.writerows(starter_rows)


def load_mapping():
    """Load keyword -> (expense_type, merchant_category, store_name) mapping."""
    mapping = []
    if not os.path.exists(MAPPING_FILE):
        return mapping
    with open(MAPPING_FILE, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            # Normalize keys (handle BOM / stray spaces)
            normalized = {k.strip().lstrip("\ufeff"): v for k, v in row.items()}
            keyword = (normalized.get("Keyword Pattern") or normalized.get("Keyword") or "").strip()
            expense_type = (normalized.get("Expense Type") or normalized.get("Category") or "").strip()
            merchant_category = (normalized.get("Merchant Category") or "").strip()
            store_name = (normalized.get("Store Name") or "").strip()
            if keyword:
                mapping.append((keyword.upper(), expense_type, merchant_category, store_name))
    return mapping


def categorize(description, mapping):
    """Assign Expense Type, Merchant Category, Store Name based on mapping."""
    desc = (description or "").upper()
    for keyword, expense_type, merchant_category, store_name in mapping:
        if keyword in desc:
            if not expense_type:
                expense_type = infer_expense_type(desc)
            if not merchant_category:
                merchant_category = "Uncategorized"
            if not store_name:
                store_name = keyword.title()
            return expense_type, merchant_category, store_name
    # No mapping found
    return infer_expense_type(desc), "Uncategorized", "Unknown"


def infer_expense_type(desc_upper):
    """Infer Expense Type from description keywords."""
    for category, keywords in EXPENSE_TYPE_KEYWORDS.items():
        for kw in keywords:
            if kw in desc_upper:
                return category
    return "Uncategorized"


def is_payment(description, expense_type, merchant_category):
    """Identify credit card bill payments."""
    if expense_type == "Card Payment" or merchant_category == "CC Bill Payment":
        return True
    desc = (description or "").upper()
    for kw in PAYMENT_KEYWORDS:
        if kw in desc:
            return True
    return False


def extract_statement_due(pdf_path):
    """Extract total payment due / outstanding from a statement PDF."""
    def parse_text(text_norm):
        # Prefer extracting from summary blocks if present
        start_labels = [
            r"PAYMENT SUMMARY",
            r"STATEMENT SUMMARY",
            r"SUMMARY AS BILLED",
            r"STATEMENT AT A GLANCE",
            r"THIS MONTH'S STATEMENT AT A GLANCE",
        ]
        end_labels = [
            r"ACCOUNT SUMMARY",
            r"TRANSACTION DETAILS",
            r"CREDIT SUMMARY",
            r"SPENDS OVERVIEW",
        ]

        block_text = text_norm
        candidates = []
        for start_label in start_labels:
            m = re.search(start_label, text_norm, flags=re.IGNORECASE)
            if not m:
                continue
            start = m.start()
            end = len(text_norm)
            end_pos = None
            for end_label in end_labels:
                m2 = re.search(end_label, text_norm[start:], flags=re.IGNORECASE)
                if m2:
                    pos = start + m2.start()
                    if end_pos is None or pos < end_pos:
                        end_pos = pos
            if end_pos is not None:
                end = end_pos
            candidates.append((start, end))

        if candidates:
            start, end = sorted(candidates, key=lambda x: x[0])[0]
            block_text = text_norm[start:end]

        label_patterns = [
            r"Total Payment Due",
            r"Total Amount Due",
            r"Total Amount due",
            r"Billed Amount",
            r"Purchases\s*/\s*Debits",
            r"Purchases/ Debits",
            r"Purchases/Debits",
        ]

        # If labels not found in block, fall back to full text
        if not any(re.search(pat, block_text, flags=re.IGNORECASE) for pat in label_patterns):
            block_text = text_norm

        for label in label_patterns:
            for match in re.finditer(label, block_text, flags=re.IGNORECASE):
                # Limit snippet to before the next label to avoid capturing unrelated amounts
                next_positions = []
                for nxt_label in label_patterns:
                    m_next = re.search(nxt_label, block_text[match.end():], flags=re.IGNORECASE)
                    if m_next:
                        next_positions.append(match.end() + m_next.start())
                next_pos = min(next_positions) if next_positions else None
                snippet = block_text[match.end():next_pos] if next_pos else block_text[match.end():match.end() + 200]
                stop_tokens = [
                    "CREDIT LIMIT",
                    "AVAILABLE CREDIT",
                    "CASH LIMIT",
                    "PAYMENT DUE DATE",
                    "MINIMUM PAYMENT DUE",
                    "STATEMENT PERIOD",
                    "DUE DATE",
                    "PREVIOUS BALANCE",
                    "ACCOUNT SUMMARY",
                ]
                # Only cut at stop tokens if they appear after the first amount
                first_amount_idx = None
                for m_amt in re.finditer(r"[`₹rR]?\s*[0-9][0-9,]*(?:\.\d{2})?", snippet):
                    first_amount_idx = m_amt.start()
                    break
                if first_amount_idx is not None:
                    stop_idx = None
                    for token in stop_tokens:
                        m_stop = re.search(token, snippet, flags=re.IGNORECASE)
                        if m_stop and m_stop.start() > first_amount_idx:
                            idx = m_stop.start()
                            if stop_idx is None or idx < stop_idx:
                                stop_idx = idx
                    if stop_idx is not None:
                        snippet = snippet[:stop_idx]

                amounts = list(re.finditer(r"[`₹rR]?\s*([0-9][0-9,]*)(?:\.(\d{2}))?\s*(Dr|CR|Cr|DR)?", snippet))
                vals = []
                vals_with_dr = []
                vals_with_dr_pos = []
                for m_amt in amounts:
                    amt, dec, drcr = m_amt.groups()
                    try:
                        val_str = amt.replace(",", "")
                        if dec:
                            val_str = f"{val_str}.{dec}"
                        val = float(val_str)
                        if val < 100:
                            continue
                        # Skip likely dates (e.g., 2026) if surrounded by slashes
                        ctx = snippet[max(0, m_amt.start() - 2):m_amt.start() + 2]
                        if "/" in ctx and 1900 <= val <= 2100:
                            continue
                        context = snippet[max(0, m_amt.start() - 40):m_amt.start()]
                        if re.search(r"CREDIT LIMIT", context, flags=re.IGNORECASE):
                            continue
                        vals.append(val)
                        if drcr:
                            vals_with_dr.append(val)
                            vals_with_dr_pos.append((val, m_amt.start()))
                    except Exception:
                        continue
                if vals_with_dr:
                    # For Total Payment Due, prefer the last DR/CR value (Axis OCR ordering)
                    if re.search(r"Total Payment Due", label, flags=re.IGNORECASE):
                        if re.search(r"Minimum Payment Due", snippet, flags=re.IGNORECASE):
                            return vals_with_dr[0]
                        return vals_with_dr[-1]
                    # If snippet contains Account Summary, prefer the last DR/CR value
                    if re.search(r"ACCOUNT SUMMARY", snippet, flags=re.IGNORECASE):
                        return vals_with_dr[-1]
                    return vals_with_dr[0]
                if vals:
                    return vals[0]
        return None

    try:
        with pdfplumber.open(pdf_path) as pdf:
            text = "\n".join([(p.extract_text() or "") for p in pdf.pages])
    except Exception:
        return None

    text_norm = re.sub(r"\s+", " ", text)
    val = parse_text(text_norm)
    if val is not None:
        return val

    # OCR fallback
    try:
        from pdf2image import convert_from_path
        import pytesseract

        images = convert_from_path(pdf_path)
        text = "\n".join(pytesseract.image_to_string(img, lang="eng") for img in images)
        text_norm = re.sub(r"\s+", " ", text)
        return parse_text(text_norm)
    except Exception:
        return None


def sort_key(record):
    """Sort by Account, Period, Date"""
    account = record.get("Account", "")
    period = record.get("Period", "")
    date_str = record.get("Date", "")
    try:
        date_obj = datetime.strptime(date_str, "%d/%m/%Y")
    except Exception:
        date_obj = datetime.min
    return (account, period, date_obj)


def aggregate():
    """
    Main aggregation function
    """
    print("\n" + "="*70)
    print("MASTER CREDIT CARD AGGREGATOR")
    print("="*70)
    print(f"Scanning: {BASE_DIR}\n")

    all_records = []
    statement_due_map = {}
    stats = {
        "total": 0,
        "success": 0,
        "failed": 0
    }

    ensure_mapping_file()
    mapping = load_mapping()

    for root, dirs, files in os.walk(BASE_DIR):
        for file in files:
            if not file.lower().endswith(".pdf"):
                continue

            file_path = os.path.join(root, file)
            stats["total"] += 1

            print(f"📄 Processing: {file}")

            parser, bank = get_parser(file_path)

            if not parser:
                print(f"   ⚠️ No parser found")
                stats["failed"] += 1
                continue

            try:
                records = parser(file_path)

                if not records:
                    print(f"   ⚠️ No transactions extracted")
                    stats["failed"] += 1
                    continue

                records = normalize(records)

                # Categorize each record
                for r in records:
                    expense_type, merchant_category, store_name = categorize(
                        r.get("Description", ""), mapping
                    )
                    # Override for Uni Gold UPI expenses
                    if "UNI GOLD CARD UPI" in str(r.get("Account", "")).upper():
                        expense_type = "Personal"
                        merchant_category = "Leisure"
                        store_name = "UPI"
                    r["Expense Type"] = expense_type
                    r["Merchant Category"] = merchant_category
                    r["Store Name"] = store_name

                # Capture statement due for reconciliation
                statement_due = extract_statement_due(file_path)
                if statement_due is not None and records:
                    account = records[0].get("Account", bank or "")
                    period = records[0].get("Period", "Unknown")
                    key = (account, period)
                    existing = statement_due_map.get(key)
                    if existing is None or statement_due > existing:
                        statement_due_map[key] = statement_due

                print(f"   ✅ Extracted {len(records)} transactions (Period: {records[0].get('Period', 'Unknown')})")

                all_records.extend(records)
                stats["success"] += 1

            except Exception as e:
                print(f"   ❌ Error: {e}")
                stats["failed"] += 1
                import traceback
                traceback.print_exc()

    # Split expenses vs payments
    expenses = []
    payments = []

    for r in all_records:
        if is_payment(
            r.get("Description", ""),
            r.get("Expense Type", ""),
            r.get("Merchant Category", ""),
        ):
            payments.append(r)
        else:
            expenses.append(r)

    # Ensure a payment row exists per Account+Period when expenses exist
    expense_keys = {(r.get("Account"), r.get("Period")) for r in expenses}
    payment_keys = {(r.get("Account"), r.get("Period")) for r in payments}
    for key in expense_keys - payment_keys:
        account, period = key
        payments.append(
            {
                "Period": period,
                "Account": account,
                "Date": "",
                "Description": "No outstanding",
                "Amount": 0.0,
                "Type": "",
            }
        )

    # Write output
    print("\n" + "="*70)
    print("Writing final Excel output...")

    expenses_sorted = sorted(expenses, key=sort_key)
    payments_sorted = sorted(payments, key=sort_key)
    all_sorted = sorted(all_records, key=sort_key)

    df_expenses = pd.DataFrame(expenses_sorted)
    df_payments = pd.DataFrame(payments_sorted)
    df_summary = pd.DataFrame(all_sorted)

    if not df_summary.empty:
        summary = (
            df_summary.groupby(
                ["Expense Type", "Merchant Category", "Store Name"], dropna=False
            )["Amount"]
            .agg(TotalAmount="sum", TransactionCount="count")
            .reset_index()
        )
    else:
        summary = pd.DataFrame(
            columns=[
                "Expense Type",
                "Merchant Category",
                "Store Name",
                "TotalAmount",
                "TransactionCount",
            ]
        )

    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
        df_expenses = df_expenses.drop(
            columns=["Type", "Brand", "Expense_Type"], errors="ignore"
        )
        df_expenses.to_excel(writer, sheet_name="Credit card expenses", index=False)

        # Reconcile using expenses total vs statement due
        if not df_payments.empty:
            expense_totals = (
                df_expenses.groupby(["Account", "Period"])["Amount"]
                .sum()
                .reset_index()
            )
            expense_total_map = {
                (row["Account"], row["Period"]): round(float(row["Amount"]), 2)
                for _, row in expense_totals.iterrows()
            }
            for idx, row in df_payments.iterrows():
                key = (row.get("Account"), row.get("Period"))
                due = statement_due_map.get(key)
                expense_sum = expense_total_map.get(key)
                if expense_sum is not None:
                    df_payments.at[idx, "Current Billed Amount"] = expense_sum
                if due is not None and expense_sum is not None:
                    diff = round(float(due) - float(expense_sum), 2)
                    df_payments.at[idx, "Recon Diff"] = diff
                    df_payments.at[idx, "Reconciled?"] = "Yes" if abs(diff) <= 0.01 else "No"
                else:
                    df_payments.at[idx, "Recon Diff"] = None
                    df_payments.at[idx, "Reconciled?"] = "No"

        if not df_payments.empty:
            df_payments["Current Outstanding Amt"] = df_payments.apply(
                lambda row: statement_due_map.get(
                    (row.get("Account"), row.get("Period"))
                ),
                axis=1,
            )

        df_payments = df_payments.drop(
            columns=["Type", "Expense Type", "Merchant Category", "Store Name"],
            errors="ignore",
        )
        df_payments = df_payments.rename(
            columns={"Amount": "Last Paid Amount"}
        )
        # Reorder columns for bill payments sheet
        desired_cols = [
            "Period",
            "Account",
            "Date",
            "Description",
            "Last Paid Amount",
            "Current Outstanding Amt",
            "Current Billed Amount",
            "Reconciled?",
            "Recon Diff",
        ]
        existing_cols = [c for c in desired_cols if c in df_payments.columns]
        remaining = [c for c in df_payments.columns if c not in existing_cols]
        df_payments = df_payments[existing_cols + remaining]
        df_payments.to_excel(writer, sheet_name="Credit card Reconciliation", index=False)
        summary.to_excel(writer, sheet_name="Credit card summary", index=False)

        # Format headers and borders across all sheets
        wb = writer.book
        from openpyxl.styles import Font, PatternFill, Border, Side

        header_fill = PatternFill(start_color="F4B183", end_color="F4B183", fill_type="solid")
        header_font = Font(bold=True)
        thin = Side(style="thin")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for ws in wb.worksheets:
            max_row = ws.max_row
            max_col = ws.max_column
            if max_row == 0 or max_col == 0:
                continue
            # Header styling
            for col in range(1, max_col + 1):
                cell = ws.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
            # Borders for all cells
            for row in range(1, max_row + 1):
                for col in range(1, max_col + 1):
                    ws.cell(row=row, column=col).border = border

    print("\n" + "="*70)
    print("✅ AGGREGATION COMPLETE")
    print("="*70)
    print(f"Total PDFs:             {stats['total']}")
    print(f"Successfully parsed:    {stats['success']}")
    print(f"Failed:                 {stats['failed']}")
    print(f"Total Transactions:     {len(all_records)}")
    print(f"Output File:            {OUTPUT_FILE}")
    print(f"Category Mapping File:  {MAPPING_FILE}")
    print("="*70 + "\n")


if __name__ == "__main__":
    aggregate()
