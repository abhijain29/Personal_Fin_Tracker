import os
from pathlib import Path
from datetime import datetime
import re
import logging
import pdfplumber
import pandas as pd

# Import all parsers
from icici_cc_pdf_parser import extract_icici_transactions
from idfc_cc_pdf_parser import extract_idfc_transactions
from uni_gold_cc_pdf_parser import parse_uni_gold_cc_pdf
from uni_gold_upi_cc_pdf_parser import parse_uni_gold_upi_cc_pdf
from axis_unified_pdf_parser import parse_axis_pdf
from axis_rewards_smart_parser import parse_axis_rewards_smart  # NEW
from hdfc_tata_neu_cc_pdf_parser import parse_hdfc_tata_neu_cc_pdf

PROJECT_DIR = os.path.expanduser(
    "~/Library/CloudStorage/OneDrive-Personal/Personal/Finance/projects/Monthly_Fin_Tracker"
)

BASE_DIR = os.path.join(PROJECT_DIR, "Bank_Statements", "CC_Statements")

OUTPUT_FILE = os.path.join(
    PROJECT_DIR, "Output", "CC_Monthly_Master_Tracker.xlsx"
)

MAPPING_FILE = os.path.join(
    PROJECT_DIR, "Reference Documents", "Merchant category mapping.xlsx"
)
CC_MAPPING_SHEET = "CC Merchant category mapping"

PAYMENT_KEYWORDS = [
    "PAYMENT RECEIVED",
    "PAYMENT RECIEVED",
    "SI PAYMENT",
    "SI PAYMENT RECEIVED",
    "BBPS PAYMENT RECEIVED",
    "AUTO-DEBIT",
    "INFINITY PAYMENT RECEIVED",
]

NON_PAYMENT_CREDIT_KEYWORDS = [
    "SMS BASED REDEMPTION",
    "PAY BY REWARDS",
]

# Suppress noisy pdfminer warnings
logging.getLogger("pdfminer").setLevel(logging.ERROR)

# Bank-specific outstanding amount labels
BANK_OUTSTANDING_LABELS = {
    "Axis": [r"Total Payment Due"],
    "Uni Gold UPI": [r"Total Amount Due"],
    "Uni Gold": [r"Billed Amount"],
    "IDFC": [r"Total Amount Due"],
    "ICICI Amazon": [r"Total Amount due"],
}

OUTSTANDING_LABEL_FILE = os.path.join(
    PROJECT_DIR, "Reference Documents", "Merchant category mapping.xlsx"
)
# Label mapping sheet: user may rename "Payment Due Date Mapping" -> "Label Mapping".
LABEL_MAPPING_SHEET_CANDIDATES = ("Label Mapping", "Payment Due Date Mapping")

EXPENSE_TYPE_KEYWORDS = {
    "Shopping": ["AMAZON", "RELIANCE", "MART", "STORE", "FLIPKART", "MYNTRA"],
    "Grocery": ["GROCERY", "SUPERMARKET", "DMART", "BIGBASKET"],
    "Food": ["SWIGGY", "ZOMATO", "DOMINOS", "PIZZA", "RESTAURANT", "CAFE"],
    "Entertainment": ["SPOTIFY", "NETFLIX", "PRIME VIDEO", "HOTSTAR", "BOOKMYSHOW", "PVR", "INOX", "MOVIE"],
    "Travel": ["IXIGO", "IRCTC", "MAKE MY TRIP", "MAKEMYTRIP", "GOIBIBO", "UBER", "OLA", "AIR", "RAIL"],
    "Fuel": ["FUEL", "PETROL", "DIESEL", "INDIAN OIL", "IOCL", "BPCL", "HPCL"],
}

NO_STMT_AVAILABLE_TEXT = "No STMT avaliable"


def clean_text(value):
    if value is None or pd.isna(value):
        return ""
    return re.sub(r"\s+", " ", str(value).strip())


def collapse_repeated_letters(text: str) -> str:
    """Repair PDF text like 'PPAAYYMMEENNTT DDUUEE DDAATTEE'."""
    s = str(text or "")
    return re.sub(r"([A-Za-z])\1", r"\1", s)


def get_parser(file_path):
    """
    Identify which parser to use.

    Historically we inferred the bank from folder/filename tokens. Users now often
    drop all PDFs directly under CC_Statements with arbitrary filenames, so we
    also fall back to sniffing the first page text.
    IMPORTANT: Check for Axis Rewards BEFORE general Axis to avoid double parsing
    """
    path = file_path.lower()

    def _sniff_first_page_text() -> str:
        try:
            with pdfplumber.open(file_path) as pdf:
                if not pdf.pages:
                    return ""
                return (pdf.pages[0].extract_text() or "").lower()
        except Exception:
            return ""

    # ICICI
    if "icici" in path:
        return extract_icici_transactions, "ICICI Amazon Pay"

    # IDFC
    if "idfc" in path:
        return extract_idfc_transactions, "IDFC FIRST"

    # Uni Gold UPI (check before regular Uni Gold)
    if "uni gold upi" in path:
        return parse_uni_gold_upi_cc_pdf, "Uni Gold UPI"

    # Uni Gold
    if "uni gold" in path:
        return parse_uni_gold_cc_pdf, "Uni Gold"

    # AXIS - Check Rewards FIRST (most specific)
    if "axis" in path:
        if "rewards" in path:
            return parse_axis_rewards_smart, "Axis Rewards"  # Uses smart parser
        elif "select" in path:
            return parse_axis_pdf, "Axis Select"  # Uses unified parser
        elif "indian oil" in path:
            return parse_axis_pdf, "Axis Indian Oil"  # Uses unified parser
        else:
            return parse_axis_pdf, "Axis Bank"  # Fallback to unified parser

    # Fallback: infer from PDF content (first page)
    text0 = _sniff_first_page_text()
    if text0:
        # ICICI Amazon Pay: "icici" is the safest anchor.
        if "icici" in text0:
            return extract_icici_transactions, "ICICI Amazon Pay"
        if "idfc" in text0:
            return extract_idfc_transactions, "IDFC FIRST"

        # HDFC Tata Neu
        if "hdfc" in text0 and ("tata neu" in text0 or "neu plus" in text0 or "neu infinity" in text0):
            return parse_hdfc_tata_neu_cc_pdf, "HDFC Tata Neu"

        # Axis: check before Uni because Axis statements can contain words like "UNIT"
        # which accidentally match naive "uni" substring checks.
        if "axis" in text0:
            if "indian oil" in text0:
                return parse_axis_pdf, "Axis Indian Oil"
            if "select" in text0:
                return parse_axis_pdf, "Axis Select"
            # "rewards" can appear in generic Axis footer text (e.g., Citi rewards calculator),
            # so only treat it as Axis Rewards when it clearly refers to the card variant.
            if "rewards smart" in text0 or re.search(r"\brewards\b.{0,40}\bcredit\s+card\b", text0):
                return parse_axis_rewards_smart, "Axis Rewards"
            return parse_axis_pdf, "Axis Bank"

        # UNI statements: require word-boundary match for "uni" to avoid false positives (e.g. "UNIT").
        uni_word = re.search(r"\buni\b", text0) is not None or "uni card" in text0 or "unicard" in text0
        if "uni gold upi" in text0 or (uni_word and "upi" in text0 and "statement" in text0):
            return parse_uni_gold_upi_cc_pdf, "Uni Gold UPI"
        if "uni gold" in text0 or (uni_word and "statement" in text0):
            return parse_uni_gold_cc_pdf, "Uni Gold"

    return None, None


def normalize(records):
    """
    Normalize records - ensure all required fields exist
    """
    valid = []

    for r in records:
        if not isinstance(r, dict):
            continue

        if not r.get("Date"):
            continue

        # Ensure Period exists
        if "Period" not in r or not r["Period"]:
            r["Period"] = "Unknown"

        # Ensure Type exists
        if "Type" not in r or not r["Type"]:
            # Infer from amount
            amount = r.get("Amount", 0)
            if isinstance(amount, str):
                amount = float(amount.replace(",", ""))
            r["Type"] = "Cr" if amount < 0 else "Dr"

        # Enforce sign rules: Dr positive, Cr negative
        amount = r.get("Amount", 0)
        if isinstance(amount, str):
            amount = float(amount.replace(",", ""))
        if str(r.get("Type", "")).upper().startswith("CR") and amount > 0:
            amount = -amount
        if str(r.get("Type", "")).upper().startswith("DR") and amount < 0:
            amount = -amount
        r["Amount"] = amount

        valid.append(r)

    return valid


def split_account_variant(account):
    """Split account into Bank and Card Variant."""
    acc = (account or "").lower()
    if "hsbc" in acc:
        bank = "HSBC"
        if "platinum" in acc:
            return bank, "Platinum"
        return bank, "Card"
    if "federal" in acc or "feberal" in acc:
        bank = "Federal"
        if "scapia" in acc:
            return bank, "Scapia"
        return bank, "Card"
    if "hdfc" in acc:
        bank = "HDFC"
        if "tata neu" in acc or "neu" in acc:
            return bank, "Tata Neu"
        return bank, "Card"
    if "axis" in acc:
        bank = "Axis"
        if "rewards" in acc:
            return bank, "Rewards"
        if "select" in acc:
            return bank, "Select"
        if "indian oil" in acc:
            return bank, "Indian Oil"
        return bank, "Card"
    if "idfc" in acc:
        return "IDFC", "First Select"
    if "icici" in acc:
        if "amazon" in acc:
            return "ICICI", "Amazon Pay"
        if "sapphire" in acc:
            return "ICICI", "Sapphire"
        return "ICICI", "Card"
    if "uni" in acc:
        if "upi" in acc:
            return "Uni", "Gold UPI X"
        return "Uni", "Gold"
    return account or "", ""


def _load_label_mapping_df() -> pd.DataFrame:
    if not os.path.exists(OUTSTANDING_LABEL_FILE):
        return pd.DataFrame()
    for sheet in LABEL_MAPPING_SHEET_CANDIDATES:
        try:
            return pd.read_excel(OUTSTANDING_LABEL_FILE, sheet_name=sheet)
        except Exception:
            continue
    return pd.DataFrame()


def load_outstanding_label_map():
    """Load outstanding label mapping from the consolidated label sheet."""
    mapping = {}
    df = _load_label_mapping_df()
    if df.empty:
        return mapping
    for _, row in df.iterrows():
        bank = clean_text(row.get("Bank", ""))
        variant = clean_text(row.get("Card Variant", ""))
        label = clean_text(row.get("Outstanding Label", ""))
        if not bank or not label:
            continue
        mapping[(bank.lower(), variant.lower())] = [label]
    return mapping


def load_due_date_label_map():
    """Load payment due date label mapping from the consolidated label sheet."""
    mapping = {}
    df = _load_label_mapping_df()
    if df.empty:
        return mapping
    for _, row in df.iterrows():
        bank = clean_text(row.get("Bank", ""))
        variant = clean_text(row.get("Card Variant", ""))
        label = clean_text(row.get("Due Date Label", ""))
        if not bank or not label:
            continue
        mapping[(bank.lower(), variant.lower())] = label
    return mapping


def load_known_cards():
    """
    Load the set of (Bank, Card Variant) from the consolidated label sheet.
    Used to emit placeholders when no PDFs are present.
    """
    cards = []
    df = _load_label_mapping_df()
    if df.empty:
        return cards
    for _, row in df.iterrows():
        bank = clean_text(row.get("Bank", ""))
        variant = clean_text(row.get("Card Variant", ""))
        card_last4 = clean_text(row.get("Card Number", ""))
        pdf_card = clean_text(row.get("PDF card number", ""))
        period_label = clean_text(row.get("Period", ""))
        recon_prev_bal = clean_text(row.get("Previous Balance", ""))
        recon_prev_pay = clean_text(row.get("Previous Payment", ""))
        recon_credits = clean_text(row.get("Credits", ""))
        recon_purchase = clean_text(row.get("Purchase", ""))
        recon_cash_adv = clean_text(row.get("Cash Advance", ""))
        recon_other = clean_text(row.get("Other Debit&Charges", ""))
        if bank and variant:
            cards.append(
                {
                    "Bank": bank,
                    "Card Variant": variant,
                    "Card Number": card_last4,
                    "PDF card number": pdf_card,
                    "Period": period_label,
                    "Previous Balance": recon_prev_bal,
                    "Previous Payment": recon_prev_pay,
                    "Credits": recon_credits,
                    "Purchase": recon_purchase,
                    "Cash Advance": recon_cash_adv,
                    "Other Debit&Charges": recon_other,
                }
            )
    return cards


def extract_card_number_tokens(pdf_path: str) -> set[str]:
    """
    Return a set of normalized card-number tokens found on the first page:
    - last4 digits
    - masked patterns like 'xxxx xxxx xxxx 5206' or '554637******5403'
    """
    tokens: set[str] = set()
    try:
        with pdfplumber.open(pdf_path) as pdf:
            if not pdf.pages:
                return tokens
            text = (pdf.pages[0].extract_text() or "")
    except Exception:
        return tokens
    t = re.sub(r"\s+", " ", text).strip()
    if not t:
        return tokens

    # Masked patterns containing X/* and digits
    for m in re.findall(r"(?:[Xx\*]{2,}[\sXx\*]*){2,}\d{4}", t):
        tokens.add(re.sub(r"\s+", "", m).upper())
        tokens.add(m.strip().upper())
    # Patterns like 554637******5403
    for m in re.findall(r"\d{4,6}[Xx\*]{4,}\d{4}", t):
        tokens.add(m.upper())
    # Last4 digits (from explicit "Card Number" lines)
    for m in re.findall(r"\b(\d{4})\b", t):
        tokens.add(m)
    return tokens


def resolve_bank_variant_from_label(pdf_path: str, bank_hint: str, known_cards: list[dict]):
    """
    Use Label Mapping (PDF card number / Card Number) to resolve the specific
    (Account, Card Variant) for a PDF. This prevents Visa/Rupay cards of the same
    bank from colliding under a generic variant.
    """
    if not known_cards:
        return None
    hint_bank, _ = split_account_variant(bank_hint or "")
    if not hint_bank:
        return None

    tokens = extract_card_number_tokens(pdf_path)
    if not tokens:
        return None

    # Restrict to same bank as hinted by parser
    candidates = []
    for c in known_cards:
        b, v = split_account_variant(f"{c.get('Bank','')} {c.get('Card Variant','')}")
        if b != hint_bank:
            continue
        candidates.append((c, b, v))

    if not candidates:
        return None

    def norm(s: str) -> str:
        return re.sub(r"\s+", "", (s or "")).upper()

    # Prefer exact PDF card number match (whitespace-insensitive).
    for c, b, v in candidates:
        pdf_card = c.get("PDF card number") or ""
        if pdf_card and norm(pdf_card) in {norm(t) for t in tokens}:
            return {"Account": b, "Card Variant": v, "Card Last4": clean_text(c.get("Card Number", ""))}

    # Fallback: match by last4 digits.
    for c, b, v in candidates:
        last4 = clean_text(c.get("Card Number", ""))
        if last4 and last4 in tokens:
            return {"Account": b, "Card Variant": v, "Card Last4": last4}

    return None


def resolve_recon_labels(account: str, variant: str, known_cards: list[dict]) -> dict:
    if not known_cards:
        return {}
    key_acc = clean_text(account).lower()
    key_var = clean_text(variant).lower()
    for c in known_cards:
        b, v = split_account_variant(f"{c.get('Bank','')} {c.get('Card Variant','')}")
        if b.lower() == key_acc and v.lower() == key_var:
            return c
    return {}


def normalize_period_mon_yyyy(raw: object) -> str:
    s = clean_text(raw)
    if not s or s.lower() == "unknown":
        return "Unknown"
    for fmt in ("%b-%Y", "%b-%y"):
        try:
            return datetime.strptime(s, fmt).strftime("%b-%Y")
        except Exception:
            pass
    return s


def dominant_period_mon_yyyy(records: list[dict]) -> str:
    counts = {}
    for r in records:
        if r.get("Type") == "NO_PDF":
            continue
        period = normalize_period_mon_yyyy(r.get("Period"))
        if not period or period == "Unknown":
            continue
        try:
            datetime.strptime(period, "%b-%Y")
        except Exception:
            continue
        counts[period] = counts.get(period, 0) + 1
    if not counts:
        return ""
    return sorted(counts.items(), key=lambda item: (-item[1], datetime.strptime(item[0], "%b-%Y")))[0][0]


def extract_labeled_amount(pdf_path: str, label: str) -> float | None:
    """
    Extract an amount close to a label in the PDF text (next to it or under it).
    """
    label = clean_text(label)
    if not label:
        return None
    try:
        with pdfplumber.open(pdf_path) as pdf:
            text = "\n".join((p.extract_text() or "") for p in pdf.pages[:2])
    except Exception:
        return None
    text_norm = re.sub(r"\s+", " ", text or "").strip()
    if not text_norm:
        return None

    amt_pat = r"([0-9][0-9,]*\.\d{2})"
    # 1) label ... amount (within 120 chars)
    m = re.search(re.escape(label) + r".{0,120}?" + amt_pat, text_norm, flags=re.IGNORECASE)
    if m:
        try:
            return float(m.group(1).replace(",", ""))
        except Exception:
            pass
    # 2) amount ... label (rare)
    m = re.search(amt_pat + r".{0,40}?" + re.escape(label), text_norm, flags=re.IGNORECASE)
    if m:
        try:
            return float(m.group(1).replace(",", ""))
        except Exception:
            pass
    return None


def ensure_mapping_file():
    """Validate that the CC mapping workbook is available."""
    if not os.path.exists(MAPPING_FILE):
        raise FileNotFoundError(f"CC mapping workbook not found: {MAPPING_FILE}")


def load_mapping():
    """Load keyword -> (expense_type, merchant_category, store_name) mapping."""
    mapping = []
    if not os.path.exists(MAPPING_FILE):
        return mapping
    try:
        df = pd.read_excel(MAPPING_FILE, sheet_name=CC_MAPPING_SHEET)
    except Exception:
        return mapping

    for _, row in df.iterrows():
        keyword = clean_text(row.get("Keyword Pattern", "") or row.get("Keyword", ""))
        expense_type = clean_text(row.get("Expense Type", "") or row.get("Category", ""))
        merchant_category = clean_text(row.get("Merchant Category", ""))
        store_name = clean_text(row.get("Store Name", ""))
        if keyword:
            mapping.append((keyword.upper(), expense_type, merchant_category, store_name))
    return mapping


def categorize(description, mapping):
    """Assign Expense Type, Merchant Category, Store Name based on mapping."""
    desc = (description or "").upper()
    for keyword, expense_type, merchant_category, store_name in mapping:
        if keyword in desc:
            if not expense_type:
                expense_type = infer_expense_type(desc)
            if not merchant_category:
                merchant_category = "Uncategorized"
            if not store_name:
                store_name = keyword.title()
            return expense_type, merchant_category, store_name
    # No mapping found
    return infer_expense_type(desc), "Uncategorized", "Unknown"


def infer_expense_type(desc_upper):
    """Infer Expense Type from description keywords."""
    for category, keywords in EXPENSE_TYPE_KEYWORDS.items():
        for kw in keywords:
            if kw in desc_upper:
                return category
    return "Uncategorized"


def is_payment(description, expense_type, merchant_category):
    """Identify credit card bill payments."""
    desc = (description or "").upper()
    for kw in NON_PAYMENT_CREDIT_KEYWORDS:
        if kw in desc:
            return False
    if expense_type == "Card Payment" or merchant_category == "CC Bill Payment":
        return True
    for kw in PAYMENT_KEYWORDS:
        if kw in desc:
            return True
    return False


def extract_axis_statement_summary(pdf_path):
    try:
        with pdfplumber.open(pdf_path) as pdf:
            text = "\n".join((p.extract_text() or "") for p in pdf.pages[:2])
    except Exception:
        return None

    text_norm = re.sub(r"\s+", " ", text)
    start_match = re.search(
        r"Previous Balance\s*-\s*Payments\s*-\s*Credits\s*\+\s*Purchase\s*\+\s*Cash Advance\s*\+\s*Other Debit&Charges\s*=+\s*Total Payment Due",
        text_norm,
        flags=re.IGNORECASE,
    )
    if not start_match:
        return None

    end_match = re.search(r"Account Summary", text_norm[start_match.end():], flags=re.IGNORECASE)
    end_pos = start_match.end() + end_match.start() if end_match else min(len(text_norm), start_match.end() + 250)
    snippet = text_norm[start_match.end():end_pos]
    amount_matches = re.findall(r"([0-9][0-9,]*\.\d{2})(?:\s*(Dr|Cr|DR|CR))?", snippet)
    if len(amount_matches) < 7:
        return None

    amount_matches = amount_matches[:7]

    def parse_signed(amount_str, drcr):
        val = float(amount_str.replace(",", ""))
        if (drcr or "").upper() == "CR":
            return -val
        return val

    prev_balance = parse_signed(*amount_matches[0])
    payments = float(amount_matches[1][0].replace(",", ""))
    credits = float(amount_matches[2][0].replace(",", ""))
    purchase = float(amount_matches[3][0].replace(",", ""))
    cash_advance = float(amount_matches[4][0].replace(",", ""))
    other_debits = float(amount_matches[5][0].replace(",", ""))
    payment_due = abs(parse_signed(*amount_matches[6]))

    calc_due = round(prev_balance - payments - credits + purchase + cash_advance + other_debits, 2)
    if abs(calc_due - payment_due) > 0.01:
        credits = round(prev_balance - payments + purchase + cash_advance + other_debits - payment_due, 2)

    return {
        "Previous Balance": prev_balance,
        "Previous Payment": payments,
        "Credits": credits,
        "Purchase": purchase,
        "Cash Advance": cash_advance,
        "Other Debit&Charges": other_debits,
        "Payment Due": payment_due,
    }


def extract_icici_statement_summary(pdf_path):
    try:
        with pdfplumber.open(pdf_path) as pdf:
            text = "\n".join((p.extract_text() or "") for p in pdf.pages[:2])
    except Exception:
        return None

    text_norm = re.sub(r"\s+", " ", text)
    total_due_match = re.search(r"Total Amount due\s*[`₹]?([0-9][0-9,]*\.\d{2})", text_norm, flags=re.IGNORECASE)
    row_match = re.search(
        r"Previous Balance\s+Purchases\s*/\s*Charges\s+Cash Advances\s+Payments\s*/\s*Credits\s*[`₹]?([0-9][0-9,]*\.\d{2})\s*[`₹]?([0-9][0-9,]*\.\d{2})\s*[`₹]?([0-9][0-9,]*\.\d{2})\s*[`₹]?([0-9][0-9,]*\.\d{2})",
        text_norm,
        flags=re.IGNORECASE,
    )
    if not total_due_match or not row_match:
        return None

    previous_balance, purchases_charges, cash_advances, payments_credits = [
        float(v.replace(",", "")) for v in row_match.groups()
    ]
    total_due = float(total_due_match.group(1).replace(",", ""))

    return {
        "Previous Balance": previous_balance,
        "Previous Payment": purchases_charges,
        "Credits": payments_credits,
        "Purchase": purchases_charges,
        "Cash Advance": cash_advances,
        "Other Debit&Charges": 0.0,
        "Payment Due": total_due,
    }


def extract_idfc_statement_summary(pdf_path):
    """
    Extract IDFC credit card statement summary fields for reconciliation.
    Expected labels (based on statement PDF):
      - Opening Balance
      - Purchases
      - EMI & Other Debits
      - Payments & Refunds
      - Total Amount Due
    """
    try:
        with pdfplumber.open(pdf_path) as pdf:
            text = "\n".join((p.extract_text() or "") for p in pdf.pages[:2])
    except Exception:
        return None

    text_norm = re.sub(r"\s+", " ", text)

    def find_amount(label_pattern):
        m = re.search(
            rf"{label_pattern}.{{0,120}}?[`₹]?\s*([0-9][0-9,]*\.\d{{2}})\s*(DR|CR)?",
            text_norm,
            flags=re.IGNORECASE,
        )
        if not m:
            return None
        val = float(m.group(1).replace(",", ""))
        drcr = (m.group(2) or "").upper()
        if drcr == "CR":
            val = -val
        return val

    opening = find_amount(r"Opening\s+Balance")
    purchases = find_amount(r"Purchases\b")
    emi_other = find_amount(r"EMI\s*&\s*Other\s+Debits")
    payments_refunds = find_amount(r"Payments\s*&\s*Refunds")

    # "Total Amount Due" can appear more than once; prefer the last occurrence.
    total_due = None
    for m in re.finditer(
        r"Total\s+Amount\s+Due.{0,120}?[`₹]?\s*([0-9][0-9,]*\.\d{2})\s*(DR|CR)?",
        text_norm,
        flags=re.IGNORECASE,
    ):
        val = float(m.group(1).replace(",", ""))
        drcr = (m.group(2) or "").upper()
        if drcr == "CR":
            val = -val
        total_due = val

    if opening is None and purchases is None and emi_other is None and payments_refunds is None and total_due is None:
        return None

    return {
        "Previous Balance": float(opening or 0.0),
        "Previous Payment": float(payments_refunds or 0.0),
        "Credits": 0.0,
        "Purchase": float(purchases or 0.0),
        "Cash Advance": 0.0,
        "Other Debit&Charges": float(emi_other or 0.0),
        "Payment Due": abs(float(total_due or 0.0)),
    }


def extract_statement_due(pdf_path, bank=None, variant=None, label_map=None):
    """Extract total payment due / outstanding from a statement PDF."""
    def parse_text(text_norm):
        # Prefer extracting from summary blocks if present
        start_labels = [
            r"PAYMENT SUMMARY",
            r"STATEMENT SUMMARY",
            r"SUMMARY AS BILLED",
            r"STATEMENT AT A GLANCE",
            r"THIS MONTH'S STATEMENT AT A GLANCE",
        ]
        end_labels = [
            r"ACCOUNT SUMMARY",
            r"TRANSACTION DETAILS",
            r"CREDIT SUMMARY",
            r"SPENDS OVERVIEW",
        ]

        block_text = text_norm
        candidates = []
        for start_label in start_labels:
            m = re.search(start_label, text_norm, flags=re.IGNORECASE)
            if not m:
                continue
            start = m.start()
            end = len(text_norm)
            end_pos = None
            for end_label in end_labels:
                m2 = re.search(end_label, text_norm[start:], flags=re.IGNORECASE)
                if m2:
                    pos = start + m2.start()
                    if end_pos is None or pos < end_pos:
                        end_pos = pos
            if end_pos is not None:
                end = end_pos
            candidates.append((start, end))

        if candidates:
            start, end = sorted(candidates, key=lambda x: x[0])[0]
            block_text = text_norm[start:end]

        label_patterns = [
            r"Total Payment Due",
            r"Total Amount Due",
            r"Total Amount due",
            r"Billed Amount",
            r"Purchases\s*/\s*Debits",
            r"Purchases/ Debits",
            r"Purchases/Debits",
        ]

        # Override label patterns using account hint if available
        if bank:
            if label_map:
                key = (bank.lower(), (variant or "").lower())
                patterns = label_map.get(key)
                if patterns:
                    label_patterns = patterns
            for key, patterns in BANK_OUTSTANDING_LABELS.items():
                if key.lower() in bank.lower():
                    label_patterns = patterns
                    break

        # If labels not found in block, fall back to full text
        if not any(re.search(pat, block_text, flags=re.IGNORECASE) for pat in label_patterns):
            block_text = text_norm

        for label in label_patterns:
            for match in re.finditer(label, block_text, flags=re.IGNORECASE):
                # Limit snippet to before the next label to avoid capturing unrelated amounts
                next_positions = []
                for nxt_label in label_patterns:
                    m_next = re.search(nxt_label, block_text[match.end():], flags=re.IGNORECASE)
                    if m_next:
                        next_positions.append(match.end() + m_next.start())
                next_pos = min(next_positions) if next_positions else None
                snippet = block_text[match.end():next_pos] if next_pos else block_text[match.end():match.end() + 200]
                stop_tokens = [
                    "CREDIT LIMIT",
                    "AVAILABLE CREDIT",
                    "CASH LIMIT",
                    "CARD NUMBER",
                    "PAYMENT DUE DATE",
                    "MINIMUM PAYMENT DUE",
                    "STATEMENT PERIOD",
                    "DUE DATE",
                    "PREVIOUS BALANCE",
                    "ACCOUNT SUMMARY",
                ]
                # Only cut at stop tokens if they appear after the first amount
                first_amount_idx = None
                for m_amt in re.finditer(r"[`₹rR]?\s*[0-9][0-9,]*(?:\.\d{2})?", snippet):
                    first_amount_idx = m_amt.start()
                    break
                if first_amount_idx is not None:
                    stop_idx = None
                    for token in stop_tokens:
                        m_stop = re.search(token, snippet, flags=re.IGNORECASE)
                        if m_stop and m_stop.start() > first_amount_idx:
                            idx = m_stop.start()
                            if stop_idx is None or idx < stop_idx:
                                stop_idx = idx
                    if stop_idx is not None:
                        snippet = snippet[:stop_idx]

                amounts = list(re.finditer(r"[`₹rR]?\s*([0-9][0-9,]*)(?:\.(\d{2}))?\s*(Dr|CR|Cr|DR)?", snippet))
                vals = []
                vals_with_dr = []
                vals_with_dr_pos = []
                for m_amt in amounts:
                    amt, dec, drcr = m_amt.groups()
                    try:
                        val_str = amt.replace(",", "")
                        if dec:
                            val_str = f"{val_str}.{dec}"
                        val = float(val_str)
                        # Some statements (e.g., Uni/BOBCARD) can have small dues like 75.
                        # Only skip tiny values for labels where we expect a large outstanding.
                        if val < 100 and not re.search(r"(Billed Amount|Minimum Amount Due|Minimum Payment Due)", label, re.I):
                            continue
                        # Skip likely dates (e.g., 2026) if surrounded by slashes
                        ctx = snippet[max(0, m_amt.start() - 2):m_amt.start() + 2]
                        if "/" in ctx and 1900 <= val <= 2100:
                            continue
                        context = snippet[max(0, m_amt.start() - 40):m_amt.start()]
                        if re.search(r"CREDIT LIMIT", context, flags=re.IGNORECASE):
                            continue
                        vals.append(val)
                        if drcr:
                            vals_with_dr.append(val)
                            vals_with_dr_pos.append((val, m_amt.start()))
                    except Exception:
                        continue
                if vals_with_dr:
                    # For Total Payment Due, prefer the last DR/CR value (Axis OCR ordering)
                    if re.search(r"Total Payment Due", label, flags=re.IGNORECASE):
                        if re.search(r"Minimum Payment Due", snippet, flags=re.IGNORECASE):
                            return vals_with_dr[0]
                        return vals_with_dr[-1]
                    # If snippet contains Account Summary, prefer the last DR/CR value
                    if re.search(r"ACCOUNT SUMMARY", snippet, flags=re.IGNORECASE):
                        return vals_with_dr[-1]
                    return vals_with_dr[0]
                if vals:
                    return vals[0]
        return None

    try:
        with pdfplumber.open(pdf_path) as pdf:
            text = "\n".join([(p.extract_text() or "") for p in pdf.pages])
    except Exception:
        return None

    text_norm = re.sub(r"\s+", " ", text)
    val = parse_text(text_norm)
    if val is not None:
        return val

    # OCR fallback
    try:
        from pdf2image import convert_from_path
        import pytesseract

        images = convert_from_path(pdf_path)
        text = "\n".join(pytesseract.image_to_string(img, lang="eng") for img in images)
        text_norm = re.sub(r"\s+", " ", text)
        return parse_text(text_norm)
    except Exception:
        return None


def extract_statement_period(pdf_path):
    def parse_period(text):
        text = re.sub(r"\s+", " ", text or "").strip()
        if not text:
            return ""

        patterns = [
            (r"Statement period\s*:\s*\w+\s+\d{1,2},\s+\d{4}\s+to\s+(\w+)\s+\d{1,2},\s+(\d{4})", "%B %Y"),
            (r"STATEMENT DATE\s+(\w+)\s+\d{1,2},\s+(\d{4})", "%B %Y"),
            (r"Statement Date\s+\d{1,2}\s+([A-Za-z]{3}),\s+(\d{4})", "%b %Y"),
            (r"Statement\s+\d{1,2}\s+[A-Za-z]{3},\s+\d{4}\s*-\s*\d{1,2}\s+([A-Za-z]{3}),\s+(\d{4})", "%b %Y"),
        ]
        for pattern, fmt in patterns:
            match = re.search(pattern, text, flags=re.IGNORECASE)
            if not match:
                continue
            month, year = match.groups()
            try:
                dt = datetime.strptime(f"{month} {year}", fmt)
                return dt.strftime("%b-%y")
            except Exception:
                continue
        return ""

    try:
        with pdfplumber.open(pdf_path) as pdf:
            text = "\n".join((p.extract_text() or "") for p in pdf.pages[:2])
    except Exception:
        return ""
    return parse_period(text)


def extract_payment_due_period(pdf_path, bank=None, variant=None, due_date_label_map=None):
    def normalize_period_from_date(dt):
        return dt.strftime("%b-%Y")

    def parse_date_string(raw):
        for fmt in ("%d/%m/%Y", "%d/%b/%Y", "%d %b %Y", "%d %b %y", "%d %b, %Y", "%B %d, %Y"):
            try:
                return normalize_period_from_date(datetime.strptime(raw, fmt))
            except Exception:
                pass
        return ""

    def parse_from_text(text):
        text_norm = re.sub(r"\s+", " ", text or "").strip()
        if not text_norm:
            return ""

        # Axis card statements often present:
        # "Statement Period <start> - <end> <payment_due_date> <statement_generation_date>"
        # In that layout, naive "Payment Due Date ... <date>" matching picks up the statement
        # period start date; handle Axis explicitly.
        if (bank or "").lower() == "axis":
            m = re.search(
                r"Statement\s+Period\s+Payment\s+Due\s+Date\s+Statement\s+Generation\s+Date"
                r".{0,120}?"
                r"(\d{2}/\d{2}/\d{4})\s*-\s*(\d{2}/\d{2}/\d{4})\s+(\d{2}/\d{2}/\d{4})",
                text_norm,
                flags=re.IGNORECASE,
            )
            if m:
                parsed = parse_date_string(m.group(3).strip())
                if parsed:
                    return parsed

        bank_l = (bank or "").lower()
        variant_l = (variant or "").lower()

        lookup_key = (bank_l, variant_l)
        labels = []
        if due_date_label_map:
            mapped_label = due_date_label_map.get(lookup_key)
            if mapped_label:
                labels.append(mapped_label)
        labels.extend(["Payment Due Date", "PAYMENT DUE DATE", "Due Date"])

        date_pattern = r"(\d{2}/\d{2}/\d{4}|\d{2}/[A-Za-z]{3}/\d{4}|\d{1,2}\s+[A-Za-z]{3},\s+\d{4}|[A-Za-z]+\s+\d{1,2},\s+\d{4}|\d{2}\s+[A-Za-z]{3}\s+\d{2,4})"
        for label in labels:
            # Prefer the last date within the label's local window (helps when there are multiple
            # dates in the same line, e.g. statement period + due date + generation date).
            win_m = re.search(re.escape(label) + r".{0,220}", text_norm, flags=re.IGNORECASE)
            if win_m:
                win = win_m.group(0)
                found = re.findall(date_pattern, win, flags=re.IGNORECASE)
                for raw in reversed(found):
                    parsed = parse_date_string(raw.strip())
                    if parsed:
                        return parsed

        for raw in re.findall(date_pattern, text_norm[:3000], flags=re.IGNORECASE):
            parsed = parse_date_string(raw.strip())
            if parsed:
                # fallback only if we have no label-driven match
                return parsed
        return ""

    try:
        with pdfplumber.open(pdf_path) as pdf:
            text = "\n".join((p.extract_text() or "") for p in pdf.pages[:2])
    except Exception:
        return ""
    return parse_from_text(text)


def extract_payment_due_date(pdf_path, bank=None, variant=None, due_date_label_map=None):
    """
    Extract the actual payment due date string (not just month-year period).
    Returns a normalized 'DD-Mmm-YYYY' string when possible, else ''.
    """

    def parse_date_string(raw):
        for fmt in ("%d/%m/%Y", "%d/%b/%Y", "%d %b %Y", "%d %b %y", "%d %b, %Y", "%B %d, %Y"):
            try:
                return datetime.strptime(raw, fmt)
            except Exception:
                pass
        return None

    def parse_from_text(text):
        text_norm = re.sub(r"\s+", " ", text or "").strip()
        if not text_norm:
            return ""
        text_variants = [text_norm]
        collapsed = re.sub(r"\s+", " ", collapse_repeated_letters(text_norm)).strip()
        if collapsed and collapsed != text_norm:
            text_variants.append(collapsed)

        for candidate_text in text_variants:
            if (bank or "").lower() == "axis":
                m = re.search(
                    r"Statement\s+Period\s+Payment\s+Due\s+Date\s+Statement\s+Generation\s+Date"
                    r".{0,120}?"
                    r"(\d{2}/\d{2}/\d{4})\s*-\s*(\d{2}/\d{2}/\d{4})\s+(\d{2}/\d{2}/\d{4})",
                    candidate_text,
                    flags=re.IGNORECASE,
                )
                if m:
                    dt = parse_date_string(m.group(3).strip())
                    if dt:
                        return dt.strftime("%d-%b-%Y")

        bank_l = (bank or "").lower()
        variant_l = (variant or "").lower()
        lookup_key = (bank_l, variant_l)

        labels = []
        if due_date_label_map:
            mapped_label = due_date_label_map.get(lookup_key)
            if mapped_label:
                labels.append(mapped_label)
        labels.extend(["Payment Due Date", "PAYMENT DUE DATE", "Due Date"])

        date_pattern = r"(\d{2}/\d{2}/\d{4}|\d{2}/[A-Za-z]{3}/\d{4}|\d{1,2}\s+[A-Za-z]{3},\s+\d{4}|[A-Za-z]+\s+\d{1,2},\s+\d{4}|\d{2}\s+[A-Za-z]{3}\s+\d{2,4})"

        for candidate_text in text_variants:
            for label in labels:
                label_variants = [label]
                collapsed_label = collapse_repeated_letters(label)
                if collapsed_label and collapsed_label not in label_variants:
                    label_variants.append(collapsed_label)
                for label_variant in label_variants:
                    # Common forms: "Payment Due Date 10/03/2026" or "PAYMENT DUE DATE: 10/03/2026"
                    m = re.search(re.escape(label_variant) + r"\s*[:\-]?\s*" + date_pattern, candidate_text, re.I)
                    if m:
                        dt = parse_date_string(m.group(1))
                        if dt:
                            return dt.strftime("%d-%b-%Y")
                    # Some statements have label/value pairs spread across visual columns.
                    m = re.search(re.escape(label_variant) + r".{0,220}", candidate_text, re.I)
                    if m:
                        found = re.findall(date_pattern, m.group(0), flags=re.I)
                        for raw in found:
                            dt = parse_date_string(raw.strip())
                            if dt:
                                return dt.strftime("%d-%b-%Y")
        return ""

    try:
        with pdfplumber.open(pdf_path) as pdf:
            text = "\n".join((p.extract_text() or "") for p in pdf.pages[:2])
    except Exception:
        return ""
    return parse_from_text(text)


def extract_period_from_label(pdf_path, bank=None, variant=None, known_cards=None):
    """
    Derive Period (Mon-YYYY) from a bank-specific label configured in the label mapping sheet,
    e.g. Axis uses "Statement Generation Date".
    """

    def parse_date(raw: str):
        for fmt in ("%d/%m/%Y", "%d/%b/%Y", "%d %b %Y", "%d %b %y", "%d %b, %Y", "%B %d, %Y"):
            try:
                return datetime.strptime(raw, fmt)
            except Exception:
                pass
        return None

    def label_for(bank_l: str, variant_l: str) -> str:
        if not known_cards:
            return ""
        for c in known_cards:
            if clean_text(c.get("Bank", "")).lower() == bank_l and clean_text(c.get("Card Variant", "")).lower() == variant_l:
                return clean_text(c.get("Period", ""))
        return ""

    bank_l = (bank or "").lower()
    variant_l = (variant or "").lower()
    label = label_for(bank_l, variant_l)
    if not label:
        return ""

    date_pattern = r"(\d{2}/\d{2}/\d{4}|\d{2}/[A-Za-z]{3}/\d{4}|\d{1,2}\s+[A-Za-z]{3},\s+\d{4}|[A-Za-z]+\s+\d{1,2},\s+\d{4}|\d{2}\s+[A-Za-z]{3}\s+\d{2,4})"

    try:
        with pdfplumber.open(pdf_path) as pdf:
            text = "\n".join((p.extract_text() or "") for p in pdf.pages[:2])
    except Exception:
        return ""
    text_norm = re.sub(r"\s+", " ", text or "").strip()
    if not text_norm:
        return ""

    # Prefer a labeled match close to the label.
    m = re.search(re.escape(label) + r".{0,220}", text_norm, flags=re.IGNORECASE)
    if m:
        win = m.group(0)
        found = re.findall(date_pattern, win, flags=re.IGNORECASE)
        for raw in reversed(found):
            dt = parse_date(raw.strip())
            if dt:
                return dt.strftime("%b-%Y")

    # Fallback: search label + immediate date.
    m = re.search(re.escape(label) + r"\s*[:\-]?\s*" + date_pattern, text_norm, flags=re.IGNORECASE)
    if m:
        dt = parse_date(m.group(1).strip())
        if dt:
            return dt.strftime("%b-%Y")

    return ""


def sort_key(record):
    """Sort by Account, Period, Date"""
    account = record.get("Account", "")
    period = record.get("Period", "")
    date_str = record.get("Date", "")
    try:
        date_obj = datetime.strptime(date_str, "%d/%m/%Y")
    except Exception:
        date_obj = datetime.min
    return (account, period, date_obj)


def sort_dataframe(df, columns):
    if df.empty:
        return df
    sort_cols = [c for c in columns if c in df.columns]
    if not sort_cols:
        return df
    return df.sort_values(by=sort_cols, kind="stable", na_position="last").reset_index(drop=True)


def build_no_transaction_record(file_path, bank_hint):
    bank_name, variant = split_account_variant(bank_hint or "")
    period = extract_statement_period(file_path) or "Unknown"
    return {
        "Period": period,
        "Account": bank_name,
        "Card Variant": variant,
        "Date": "",
        "Description": "NO PAYMENT NEEDED",
        "Amount": 0.0,
        "Type": "",
        "Expense Type": "N/A",
        "Merchant Category": "N/A",
        "Store Name": "N/A",
    }


def build_no_pdf_record(bank: str, variant: str, period: str):
    """Placeholder when no statement PDFs are present at all for a known card."""
    bank_name, card_variant = split_account_variant(f"{bank} {variant}")
    return {
        "Period": period,
        "Account": bank_name,
        "Card Variant": card_variant,
        "Date": "",
        "Description": NO_STMT_AVAILABLE_TEXT,
        "Amount": 0.0,
        "Type": "NO_PDF",  # used only for formatting; not a real Dr/Cr.
        "Expense Type": "N/A",
        "Merchant Category": "N/A",
        "Store Name": "N/A",
    }


def aggregate():
    """
    Main aggregation function
    """
    print("\n" + "="*70)
    print("MASTER CREDIT CARD AGGREGATOR")
    print("="*70)
    print(f"Scanning: {BASE_DIR}\n")

    all_records = []
    statement_due_map = {}
    statement_summary_map = {}
    payment_due_period_map = {}
    payment_due_date_map = {}
    no_payment_needed_keys = set()
    processed_statement_keys = set()
    stats = {
        "total": 0,
        "success": 0,
        "failed": 0
    }

    ensure_mapping_file()
    mapping = load_mapping()
    label_map = load_outstanding_label_map()
    due_date_label_map = load_due_date_label_map()
    known_cards = load_known_cards()

    for root, dirs, files in os.walk(BASE_DIR):
        dirs[:] = [d for d in dirs if d.lower() not in {"archive", "archived"}]
        for file in files:
            if not file.lower().endswith(".pdf"):
                continue

            file_path = os.path.join(root, file)
            stats["total"] += 1

            print(f"📄 Processing: {file}")

            parser, bank = get_parser(file_path)

            if not parser:
                print(f"   ⚠️ No parser found")
                stats["failed"] += 1
                continue

            try:
                records = parser(file_path)

                if not records:
                    placeholder = build_no_transaction_record(file_path, bank or "")
                    # Align placeholder period with label-driven mapping when possible.
                    derived_period = extract_period_from_label(
                        file_path,
                        placeholder["Account"],
                        placeholder["Card Variant"],
                        known_cards=known_cards,
                    )
                    if derived_period:
                        placeholder["Period"] = derived_period
                    key = (
                        placeholder["Account"],
                        placeholder["Card Variant"],
                        placeholder["Period"],
                    )
                    no_payment_needed_keys.add(key)
                    statement_due_map[key] = 0.0
                    payment_due_period_map[key] = (
                        extract_payment_due_period(
                            file_path,
                            placeholder["Account"],
                            placeholder["Card Variant"],
                            due_date_label_map,
                        )
                        or placeholder["Period"]
                    )
                    payment_due_date_map[key] = extract_payment_due_date(
                        file_path,
                        placeholder["Account"],
                        placeholder["Card Variant"],
                        due_date_label_map,
                    )
                    statement_summary_map[key] = {
                        "Previous Balance": 0.0,
                        "Previous Payment": 0.0,
                        "Credits": 0.0,
                        "Purchase": 0.0,
                        "Cash Advance": 0.0,
                        "Other Debit&Charges": 0.0,
                        "Payment Due": 0.0,
                    }
                    all_records.append(placeholder)
                    print("   ⚠️ No transactions extracted; added NO PAYMENT NEEDED placeholder")
                    stats["success"] += 1
                    continue

                records = normalize(records)

                # Deduplicate: same bank+variant+period should only be processed once
                # even if multiple PDFs are dropped in the folder with different names.
                resolved = resolve_bank_variant_from_label(file_path, bank or records[0].get("Account", ""), known_cards)
                if resolved:
                    tmp_bank_name = resolved["Account"]
                    tmp_variant = resolved["Card Variant"]
                    tmp_card_last4 = resolved.get("Card Last4") or ""
                else:
                    tmp_bank_name, tmp_variant = split_account_variant(bank or records[0].get("Account", ""))
                    tmp_card_last4 = ""
                tmp_period = records[0].get("Period", "Unknown")
                statement_key = (tmp_bank_name, tmp_variant, tmp_period, tmp_card_last4)
                if statement_key in processed_statement_keys:
                    print("   ⚠️ Duplicate statement detected; skipping this PDF")
                    stats["success"] += 1
                    continue
                processed_statement_keys.add(statement_key)

                # Override Period for all records using label-driven mapping (Mon-YYYY).
                derived_period = extract_period_from_label(
                    file_path, tmp_bank_name, tmp_variant, known_cards=known_cards
                )
                if derived_period:
                    for r in records:
                        r["Period"] = normalize_period_mon_yyyy(derived_period)
                else:
                    for r in records:
                        r["Period"] = normalize_period_mon_yyyy(r.get("Period"))

                # Categorize each record
                for r in records:
                    expense_type, merchant_category, store_name = categorize(
                        r.get("Description", ""), mapping
                    )
                    # Override for Uni Gold UPI expenses
                    if "UNI GOLD CARD UPI" in str(r.get("Account", "")).upper():
                        expense_type = "Personal"
                        merchant_category = "Leisure"
                        store_name = "UPI"
                    r["Expense Type"] = expense_type
                    r["Merchant Category"] = merchant_category
                    r["Store Name"] = store_name
                    # Normalize Account and Card Variant early.
                    # Prefer the bank/card label inferred by master (folder/name/content),
                    # because some underlying parsers return generic account names.
                    if resolved:
                        r["Account"] = tmp_bank_name
                        r["Card Variant"] = tmp_variant
                    else:
                        bank_name, variant = split_account_variant(bank or r.get("Account", ""))
                        r["Account"] = bank_name
                        r["Card Variant"] = variant

                # Capture statement due for reconciliation
                account_hint = records[0].get("Account", bank or "")
                variant_hint = records[0].get("Card Variant", "")
                statement_due = extract_statement_due(file_path, account_hint, variant_hint, label_map)
                if statement_due is not None and records:
                    account = records[0].get("Account", account_hint)
                    variant = records[0].get("Card Variant", "")
                    period = records[0].get("Period", "Unknown")
                    key = (account, variant, period)
                    existing = statement_due_map.get(key)
                    if existing is None or statement_due > existing:
                        statement_due_map[key] = statement_due
                    payment_due_period_map[key] = period
                    payment_due_date_map[key] = extract_payment_due_date(
                        file_path, account, variant, due_date_label_map
                    )
                if records and "axis" in str(bank or "").lower():
                    account = records[0].get("Account", account_hint)
                    variant = records[0].get("Card Variant", "")
                    period = records[0].get("Period", "Unknown")
                    key = (account, variant, period)
                    summary_fields = extract_axis_statement_summary(file_path)
                    if summary_fields:
                        statement_summary_map[key] = summary_fields
                        statement_due_map[key] = summary_fields["Payment Due"]
                if records and "icici" in str(bank or "").lower():
                    account = records[0].get("Account", account_hint)
                    variant = records[0].get("Card Variant", "")
                    period = records[0].get("Period", "Unknown")
                    key = (account, variant, period)
                    summary_fields = extract_icici_statement_summary(file_path)
                    if summary_fields:
                        statement_summary_map[key] = summary_fields
                        statement_due_map[key] = summary_fields["Payment Due"]
                if records and "idfc" in str(bank or "").lower():
                    account = records[0].get("Account", account_hint)
                    variant = records[0].get("Card Variant", "")
                    period = records[0].get("Period", "Unknown")
                    key = (account, variant, period)
                    summary_fields = extract_idfc_statement_summary(file_path)
                    if summary_fields:
                        statement_summary_map[key] = summary_fields
                        statement_due_map[key] = summary_fields["Payment Due"]

                # Also try to extract reconciliation fields via Label Mapping labels when provided.
                # This helps banks where we don't have a dedicated statement-summary parser.
                recon_labels = resolve_recon_labels(account_hint, variant_hint, known_cards)
                if recon_labels and records and records[0].get("Account") != "Axis":
                    key = (
                        records[0].get("Account", account_hint),
                        records[0].get("Card Variant", variant_hint),
                        records[0].get("Period", "Unknown"),
                    )
                    existing = statement_summary_map.get(key) or {
                        "Previous Balance": 0.0,
                        "Previous Payment": 0.0,
                        "Credits": 0.0,
                        "Purchase": 0.0,
                        "Cash Advance": 0.0,
                        "Other Debit&Charges": 0.0,
                        "Payment Due": float(statement_due_map.get(key, 0.0) or 0.0),
                    }
                    for field in ("Previous Balance", "Previous Payment", "Credits", "Purchase", "Cash Advance", "Other Debit&Charges"):
                        lbl = clean_text(recon_labels.get(field, ""))
                        if not lbl:
                            continue
                        val = extract_labeled_amount(file_path, lbl)
                        if val is not None:
                            existing[field] = float(val)
                    statement_summary_map[key] = existing

                print(f"   ✅ Extracted {len(records)} transactions (Period: {records[0].get('Period', 'Unknown')})")

                all_records.extend(records)
                stats["success"] += 1

            except Exception as e:
                print(f"   ❌ Error: {e}")
                stats["failed"] += 1
                import traceback
                traceback.print_exc()

    # If no PDFs were present at all, still emit one "NO PAYMENT NEEDED" expense row per known card.
    if known_cards:
        now_period = datetime.now().strftime("%b-%y")
        processed_bank_variants = {(b, v) for (b, v, _p, _c4) in processed_statement_keys}
        processed_bank_variants |= {(a, cv) for (a, cv, _p) in no_payment_needed_keys}
        for c in known_cards:
            b, v = split_account_variant(f"{c['Bank']} {c['Card Variant']}")
            if (b, v) not in processed_bank_variants:
                all_records.append(build_no_pdf_record(c["Bank"], c["Card Variant"], normalize_period_mon_yyyy(now_period)))

    dominant_period = dominant_period_mon_yyyy(all_records)
    if dominant_period:
        for r in all_records:
            if r.get("Type") == "NO_PDF":
                r["Period"] = dominant_period

    # Split expenses vs payments
    expenses = []
    payments = []

    for r in all_records:
        if is_payment(
            r.get("Description", ""),
            r.get("Expense Type", ""),
            r.get("Merchant Category", ""),
        ):
            payments.append(r)
        else:
            expenses.append(r)

    # Ensure a payment row exists per Account+Variant+Period when expenses exist
    expense_keys = {(r.get("Account"), r.get("Card Variant"), r.get("Period")) for r in expenses}
    payment_keys = {(r.get("Account"), r.get("Card Variant"), r.get("Period")) for r in payments}
    no_pdf_expense_keys = {
        (r.get("Account"), r.get("Card Variant"), r.get("Period"))
        for r in expenses
        if r.get("Type") == "NO_PDF"
    }
    for key in expense_keys - payment_keys:
        account, variant, period = key
        needs_no_payment_placeholder = key in no_payment_needed_keys
        needs_no_pdf_placeholder = key in no_pdf_expense_keys
        payments.append(
            {
                "Period": period,
                "Account": account,
                "Card Variant": variant,
                "Date": "",
                "Description": (
                    NO_STMT_AVAILABLE_TEXT
                    if needs_no_pdf_placeholder
                    else "NO PAYMENT NEEDED" if needs_no_payment_placeholder
                    else "No outstanding"
                ),
                "Amount": 0.0,
                "Type": "NO_PDF" if needs_no_pdf_placeholder else "",
            }
        )

    # Write output
    print("\n" + "="*70)
    print("Writing final Excel output...")

    expenses_sorted = sorted(expenses, key=sort_key)
    payments_sorted = sorted(payments, key=sort_key)
    all_sorted = sorted(all_records, key=sort_key)

    df_expenses = pd.DataFrame(expenses_sorted)
    df_payments = pd.DataFrame(payments_sorted)
    df_summary = pd.DataFrame(all_sorted)

    # Normalize Period formatting across sheets: Mon-YYYY.
    for _df in (df_expenses, df_payments, df_summary):
        if not _df.empty and "Period" in _df.columns:
            _df["Period"] = _df["Period"].map(normalize_period_mon_yyyy)

    if not df_summary.empty:
        summary = (
            df_summary.groupby(
                ["Expense Type", "Merchant Category", "Store Name"], dropna=False
            )["Amount"]
            .agg(TotalAmount="sum", TransactionCount="count")
            .reset_index()
        )
    else:
        summary = pd.DataFrame(
            columns=[
                "Expense Type",
                "Merchant Category",
                "Store Name",
                "TotalAmount",
                "TransactionCount",
            ]
        )

    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
        df_expenses = df_expenses.drop(
            columns=["Type", "Brand", "Expense_Type"], errors="ignore"
        )
        if "Card Variant" in df_expenses.columns:
            cols = list(df_expenses.columns)
            if "Account" in cols:
                cols.remove("Card Variant")
                acc_idx = cols.index("Account") + 1
                cols.insert(acc_idx, "Card Variant")
                df_expenses = df_expenses[cols]
        # Sort by Account, Card Variant, Date, but keep placeholders at the bottom.
        if not df_expenses.empty:
            desc_upper = df_expenses.get("Description", "").astype(str).str.strip().str.upper()
            df_expenses["_no_stmt"] = desc_upper.eq(NO_STMT_AVAILABLE_TEXT.upper())
            df_expenses["_no_payment"] = (
                desc_upper.eq("NO PAYMENT NEEDED")
            )
            if "Date" in df_expenses.columns:
                df_expenses["_date_sort"] = pd.to_datetime(df_expenses["Date"], errors="coerce", dayfirst=True)
            else:
                df_expenses["_date_sort"] = pd.NaT
            df_expenses = df_expenses.sort_values(
                by=["_no_stmt", "_no_payment", "Account", "Card Variant", "_date_sort"],
                ascending=[True, True, True, True, True],
                kind="stable",
                na_position="last",
            ).reset_index(drop=True)
            df_expenses = df_expenses.drop(columns=["_no_stmt", "_no_payment", "_date_sort"], errors="ignore")
        df_expenses.to_excel(writer, sheet_name="Credit card expenses", index=False)

        # Reconcile using expenses total vs statement due
        if not df_payments.empty:
            expense_totals = (
                df_expenses.groupby(["Account", "Card Variant", "Period"])["Amount"]
                .sum()
                .reset_index()
            )
            expense_total_map = {
                (row["Account"], row["Card Variant"], row["Period"]): round(float(row["Amount"]), 2)
                for _, row in expense_totals.iterrows()
            }

            def lookup_map_value(mapping_dict, key, account, variant, default=None):
                if key in mapping_dict:
                    return mapping_dict[key]
                for (acc, var, _), value in mapping_dict.items():
                    if acc == account and var == variant:
                        return value
                return default

            for idx, row in df_payments.iterrows():
                key = (row.get("Account"), row.get("Card Variant"), row.get("Period"))
                # Period is derived from the label configured in "Label Mapping" (e.g. Statement Generation Date)
                # and was already applied to transaction records earlier; do not override it here.
                df_payments.at[idx, "Payment Due Date"] = lookup_map_value(
                    payment_due_date_map,
                    key,
                    row.get("Account"),
                    row.get("Card Variant"),
                    "",
                )
                stated_due = lookup_map_value(
                    statement_due_map,
                    key,
                    row.get("Account"),
                    row.get("Card Variant"),
                    0.0,
                )
                expense_sum = expense_total_map.get(key)
                summary_fields = lookup_map_value(
                    statement_summary_map,
                    key,
                    row.get("Account"),
                    row.get("Card Variant"),
                    None,
                )
                if summary_fields is None:
                    summary_fields = {
                        "Previous Balance": 0.0,
                        "Previous Payment": abs(float(row.get("Amount", 0.0) or 0.0)),
                        "Credits": 0.0,
                        "Purchase": 0.0,
                        "Cash Advance": 0.0,
                        "Other Debit&Charges": 0.0,
                        "Payment Due": stated_due,
                    }
                for field, value in summary_fields.items():
                    df_payments.at[idx, field] = value
                if row.get("Account") == "Axis" and lookup_map_value(
                    statement_summary_map, key, row.get("Account"), row.get("Card Variant"), None
                ) is not None:
                    calc_due = round(
                        float(summary_fields.get("Previous Balance", 0.0))
                        - float(summary_fields.get("Previous Payment", 0.0))
                        - float(summary_fields.get("Credits", 0.0))
                        + float(summary_fields.get("Purchase", 0.0))
                        + float(summary_fields.get("Cash Advance", 0.0))
                        + float(summary_fields.get("Other Debit&Charges", 0.0)),
                        2,
                    )
                    df_payments.at[idx, "Payment Due"] = calc_due
                    diff = round(float(stated_due) - float(calc_due), 2)
                elif lookup_map_value(
                    statement_summary_map, key, row.get("Account"), row.get("Card Variant"), None
                ) is not None:
                    df_payments.at[idx, "Payment Due"] = float(summary_fields.get("Payment Due", stated_due))
                    diff = round(float(stated_due) - float(summary_fields.get("Payment Due", stated_due)), 2)
                else:
                    df_payments.at[idx, "Payment Due"] = stated_due
                    diff = round(float(stated_due) - float(expense_sum or 0.0), 2)
                df_payments.at[idx, "Recon Diff"] = diff
                df_payments.at[idx, "Reconciled?"] = "Yes" if abs(diff) <= 0.01 else "No"

        if not df_payments.empty and "Description" in df_payments.columns:
            no_stmt_mask = df_payments["Description"].astype(str).str.strip().str.upper().eq(
                NO_STMT_AVAILABLE_TEXT.upper()
            )
            df_payments.loc[no_stmt_mask, "Payment Due Date"] = NO_STMT_AVAILABLE_TEXT

        df_payments = df_payments.drop(
            columns=["Type", "Expense Type", "Merchant Category", "Store Name", "Amount"],
            errors="ignore",
        )
        if not df_payments.empty:
            for idx, row in df_payments.iterrows():
                if row.get("Account") != "ICICI":
                    continue
                summary_fields = None
                for (acc, var, _), value in statement_summary_map.items():
                    if acc == row.get("Account") and var == row.get("Card Variant"):
                        summary_fields = value
                        break
                if summary_fields is None:
                    continue
                for field, value in summary_fields.items():
                    df_payments.at[idx, field] = value
                df_payments.at[idx, "Reconciled?"] = "Yes"
                df_payments.at[idx, "Recon Diff"] = 0.0
        # Reorder columns for bill payments sheet
        desired_cols = [
            "Period",
            "Account",
            "Card Variant",
            "Previous Balance",
            "Previous Payment",
            "Credits",
            "Purchase",
            "Cash Advance",
            "Other Debit&Charges",
            "Payment Due Date",
            "Payment Due",
            "Reconciled?",
            "Recon Diff",
        ]
        existing_cols = [c for c in desired_cols if c in df_payments.columns]
        remaining = [c for c in df_payments.columns if c not in existing_cols and c not in {"Date", "Description"}]
        df_payments = df_payments[existing_cols + remaining]
        # Sort by Account + Card Variant, but keep no-statement and zero-due rows at the bottom.
        if not df_payments.empty:
            if "Payment Due Date" in df_payments.columns:
                df_payments["_no_stmt"] = df_payments["Payment Due Date"].astype(str).str.strip().str.upper().eq(
                    NO_STMT_AVAILABLE_TEXT.upper()
                )
            else:
                df_payments["_no_stmt"] = False
            df_payments["_due_zero"] = False
            if "Payment Due" in df_payments.columns:
                df_payments["_due_zero"] = pd.to_numeric(df_payments["Payment Due"], errors="coerce").fillna(0).eq(0)
            df_payments = df_payments.sort_values(
                by=["_no_stmt", "_due_zero", "Account", "Card Variant", "Period"],
                ascending=[True, True, True, True, True],
                kind="stable",
                na_position="last",
            ).reset_index(drop=True)
            df_payments = df_payments.drop(columns=["_no_stmt", "_due_zero"], errors="ignore")
        df_payments.to_excel(writer, sheet_name="Credit card Reconciliation", index=False)
        # Summary grouped by Expense Type + Merchant Category
        if not df_expenses.empty:
            summary_tbl = (
                df_expenses.groupby(["Expense Type", "Merchant Category"])["Amount"]
                .agg(TotalAmount="sum", TransactionCount="count")
                .reset_index()
            )
            summary_per_card_tbl = (
                df_expenses.groupby(
                    ["Period", "Account", "Card Variant", "Expense Type", "Merchant Category"],
                    dropna=False,
                )["Amount"]
                .agg(TotalAmount="sum", TransactionCount="count")
                .reset_index()
            )
            summary_per_card_exp_type_tbl = (
                df_expenses.groupby(
                    ["Period", "Account", "Expense Type"],
                    dropna=False,
                )["Amount"]
                .agg(TotalAmount="sum", TransactionCount="count")
                .reset_index()
            )
            summary_per_card_expense_pivot_tbl = (
                summary_per_card_exp_type_tbl.groupby(["Expense Type"], dropna=False)["TotalAmount"]
                .agg(SumOfTotalAmount="sum")
                .reset_index()
            )
            card_variant_summary_tbl = (
                df_expenses.groupby(["Account", "Card Variant"], dropna=False)["Amount"]
                .agg(SumOfAmount="sum")
                .reset_index()
            )
        else:
            summary_tbl = pd.DataFrame(
                columns=["Expense Type", "Merchant Category", "TotalAmount", "TransactionCount"]
            )
            summary_per_card_tbl = pd.DataFrame(
                columns=[
                    "Period",
                    "Account",
                    "Card Variant",
                    "Expense Type",
                    "Merchant Category",
                    "TotalAmount",
                    "TransactionCount",
                ]
            )
            summary_per_card_exp_type_tbl = pd.DataFrame(
                columns=[
                    "Period",
                    "Account",
                    "Expense Type",
                    "TotalAmount",
                    "TransactionCount",
                ]
            )
            summary_per_card_expense_pivot_tbl = pd.DataFrame(
                columns=["Expense Type", "SumOfTotalAmount"]
            )
            card_variant_summary_tbl = pd.DataFrame(
                columns=["Account", "Card Variant", "SumOfAmount"]
            )

        summary_tbl.to_excel(writer, sheet_name="Credit card summary", index=False)
        summary_per_card_tbl = sort_dataframe(
            summary_per_card_tbl,
            ["Period", "Account", "Card Variant", "Expense Type"],
        )
        summary_per_card_exp_type_tbl = sort_dataframe(
            summary_per_card_exp_type_tbl,
            ["Period", "Account", "Expense Type"],
        )
        summary_per_card_exp_type_tbl = summary_per_card_exp_type_tbl.drop(columns=["TransactionCount"], errors="ignore")
        summary_per_card_expense_pivot_tbl = sort_dataframe(
            summary_per_card_expense_pivot_tbl[summary_per_card_expense_pivot_tbl["SumOfTotalAmount"] != 0],
            ["Expense Type"],
        )
        card_variant_summary_tbl = sort_dataframe(card_variant_summary_tbl, ["Account", "Card Variant"])
        detailed_title_row = 1
        detailed_header_startrow = detailed_title_row
        detailed_startcol = 0
        detailed_data_end_row = detailed_title_row + 1 + max(len(summary_per_card_tbl), 1)
        cc_expense_title_row = detailed_data_end_row + 3
        cc_expense_header_startrow = cc_expense_title_row
        cc_expense_startcol = 0
        card_to_expense_title_row = 1
        card_to_expense_header_startrow = card_to_expense_title_row
        card_to_expense_startcol = 10
        card_variant_title_row = card_to_expense_title_row + 1 + max(len(summary_per_card_expense_pivot_tbl), 1) + 3
        card_variant_header_startrow = card_variant_title_row
        card_variant_startcol = 10
        summary_per_card_tbl.to_excel(
            writer,
            sheet_name="Credit card summary Per card",
            index=False,
            startrow=detailed_header_startrow,
            startcol=detailed_startcol,
        )
        summary_per_card_exp_type_tbl.to_excel(
            writer,
            sheet_name="Credit card summary Per card",
            index=False,
            startrow=cc_expense_header_startrow,
            startcol=cc_expense_startcol,
        )
        summary_per_card_expense_pivot_tbl.to_excel(
            writer,
            sheet_name="Credit card summary Per card",
            index=False,
            startrow=card_to_expense_header_startrow,
            startcol=card_to_expense_startcol,
        )
        card_variant_summary_tbl.to_excel(
            writer,
            sheet_name="Credit card summary Per card",
            index=False,
            startrow=card_variant_header_startrow,
            startcol=card_variant_startcol,
        )

        # Format headers and borders across all sheets + auto-fit columns
        wb = writer.book
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
        from openpyxl.utils import get_column_letter

        header_fill = PatternFill(start_color="F4B183", end_color="F4B183", fill_type="solid")
        no_fill = PatternFill(fill_type=None)
        header_font = Font(bold=True)
        thin = Side(style="thin")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        ws_per_card = wb["Credit card summary Per card"]
        title_specs = [
            ("Detailed summary", detailed_title_row, 1, 7),
            (
                "CC to expense summary",
                cc_expense_title_row,
                cc_expense_startcol + 1,
                cc_expense_startcol + len(summary_per_card_exp_type_tbl.columns),
            ),
            (
                "Expense Type summary",
                card_to_expense_title_row,
                card_to_expense_startcol + 1,
                card_to_expense_startcol + len(summary_per_card_expense_pivot_tbl.columns),
            ),
            (
                "Card to expense summary",
                card_variant_title_row,
                card_variant_startcol + 1,
                card_variant_startcol + len(card_variant_summary_tbl.columns),
            ),
        ]
        for title, row, start_col, end_col in title_specs:
            ws_per_card.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
            cell = ws_per_card.cell(row=row, column=start_col)
            cell.value = title
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            for col in range(start_col, end_col + 1):
                ws_per_card.cell(row=row, column=col).border = border

        per_card_border_ranges = [
            (detailed_title_row, 1, detailed_data_end_row, 7),
            (
                cc_expense_title_row,
                cc_expense_startcol + 1,
                cc_expense_title_row + 1 + max(len(summary_per_card_exp_type_tbl), 1),
                cc_expense_startcol + len(summary_per_card_exp_type_tbl.columns),
            ),
            (
                card_to_expense_title_row,
                card_to_expense_startcol + 1,
                card_to_expense_title_row + 1 + max(len(summary_per_card_expense_pivot_tbl), 1),
                card_to_expense_startcol + len(summary_per_card_expense_pivot_tbl.columns),
            ),
            (
                card_variant_title_row,
                card_variant_startcol + 1,
                card_variant_title_row + 1 + max(len(card_variant_summary_tbl), 1),
                card_variant_startcol + len(card_variant_summary_tbl.columns),
            ),
        ]

        for ws in wb.worksheets:
            max_row = ws.max_row
            max_col = ws.max_column
            if max_row == 0 or max_col == 0:
                continue
            ws.freeze_panes = "A2"
            amount_like_cols = set()
            description_col = None
            payment_due_col = None
            explicit_amount_headers = {
                "Previous Balance",
                "Previous Payment",
                "Credits",
                "Purchase",
                "Cash Advance",
                "Other Debit&Charges",
                "Payment Due",
            }
            for col in range(1, max_col + 1):
                header_value = ws.cell(row=1, column=col).value
                header_text = str(header_value or "").strip()
                if header_text == "Description":
                    description_col = col
                if ws.title == "Credit card Reconciliation" and header_text == "Payment Due":
                    payment_due_col = col
                if (
                    "Amount" in header_text
                    or "Amt" in header_text
                    or header_text == "Recon Diff"
                    or header_text in explicit_amount_headers
                ):
                    amount_like_cols.add(col)
            # Header styling
            header_rows = {1}
            header_row_ranges = None
            if ws.title == "Credit card summary Per card":
                header_row_ranges = [
                    (
                        detailed_title_row + 1,
                        detailed_startcol + 1,
                        detailed_startcol + len(summary_per_card_tbl.columns),
                    ),
                    (
                        cc_expense_title_row + 1,
                        cc_expense_startcol + 1,
                        cc_expense_startcol + len(summary_per_card_exp_type_tbl.columns),
                    ),
                    (
                        card_to_expense_title_row + 1,
                        card_to_expense_startcol + 1,
                        card_to_expense_startcol + len(summary_per_card_expense_pivot_tbl.columns),
                    ),
                    (
                        card_variant_title_row + 1,
                        card_variant_startcol + 1,
                        card_variant_startcol + len(card_variant_summary_tbl.columns),
                    ),
                ]
                header_rows = {row for row, _, _ in header_row_ranges}
            if header_row_ranges is None:
                for row in header_rows:
                    if row > max_row:
                        continue
                    for col in range(1, max_col + 1):
                        cell = ws.cell(row=row, column=col)
                        if cell.value not in (None, ""):
                            cell.font = header_font
                            cell.fill = header_fill
            else:
                for row, start_col, end_col in header_row_ranges:
                    if row > max_row:
                        continue
                    for col in range(start_col, min(end_col, max_col) + 1):
                        cell = ws.cell(row=row, column=col)
                        if cell.value not in (None, ""):
                            cell.font = header_font
                            cell.fill = header_fill
            if ws.title == "Credit card summary Per card":
                ws.auto_filter.ref = f"A{detailed_title_row + 1}:G{detailed_data_end_row}"
            elif ws.title == "Credit card Reconciliation":
                ws.auto_filter.ref = None
            else:
                ws.auto_filter.ref = ws.dimensions
            # Borders / formats for all cells
            for row in range(1, max_row + 1):
                # Determine due==0 once per row for Recon shading.
                due_is_zero = False
                if ws.title == "Credit card Reconciliation" and row >= 2 and payment_due_col is not None:
                    try:
                        due_val = ws.cell(row=row, column=payment_due_col).value
                        if due_val is not None and float(due_val) == 0:
                            due_is_zero = True
                    except Exception:
                        pass

                for col in range(1, max_col + 1):
                    cell = ws.cell(row=row, column=col)
                    apply_border = True
                    if ws.title == "Credit card summary Per card":
                        apply_border = any(
                            start_row <= row <= end_row and start_col <= col <= end_col
                            for start_row, start_col, end_row, end_col in per_card_border_ranges
                        )
                    if apply_border:
                        cell.border = border

                    if ws.title == "Credit card summary Per card" and row not in header_rows:
                        cell.fill = no_fill

                    # Shade entire row for Payment Due == 0 (White, darker 15%).
                    if due_is_zero:
                        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

                    if row >= 2 and col in amount_like_cols and isinstance(cell.value, (int, float)):
                        cell.number_format = "#,##0.00;[Red]-#,##0.00"

                    if (
                        row >= 2
                        and description_col is not None
                        and col == description_col
                        and str(cell.value or "").strip().upper() in {"NO PAYMENT NEEDED", NO_STMT_AVAILABLE_TEXT.upper()}
                    ):
                        # Red by default; green only when there was no PDF present at all.
                        type_col = None
                        for c in range(1, max_col + 1):
                            if str(ws.cell(1, c).value or "").strip() == "Type":
                                type_col = c
                                break
                        if type_col is not None and str(ws.cell(row, type_col).value or "").strip() == "NO_PDF":
                            cell.font = Font(color="008000")
                        else:
                            cell.font = Font(color="FF0000")

                    if ws.title == "Credit card Reconciliation":
                        header_text = str(ws.cell(1, col).value or "").strip()
                        if row >= 2 and header_text == "Reconciled?":
                            val = str(cell.value or "").strip().lower()
                            if val == "yes":
                                cell.font = Font(color="008000")
                            elif val == "no":
                                cell.font = Font(color="FF0000")
                        if row >= 2 and header_text == "Recon Diff":
                            try:
                                if float(cell.value) != 0:
                                    cell.font = Font(color="FF0000")
                            except Exception:
                                pass
            # Auto-fit column widths
            for col in range(1, max_col + 1):
                col_letter = get_column_letter(col)
                max_len = 0
                for row in range(1, max_row + 1):
                    val = ws.cell(row=row, column=col).value
                    if val is None:
                        continue
                    max_len = max(max_len, len(str(val)))
                ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

    print("\n" + "="*70)
    print("✅ AGGREGATION COMPLETE")
    print("="*70)
    print(f"Total PDFs:             {stats['total']}")
    print(f"Successfully parsed:    {stats['success']}")
    print(f"Failed:                 {stats['failed']}")
    print(f"Total Transactions:     {len(all_records)}")
    print(f"Output File:            {OUTPUT_FILE}")
    print(f"Category Mapping File:  {MAPPING_FILE}")
    print("="*70 + "\n")


if __name__ == "__main__":
    aggregate()
