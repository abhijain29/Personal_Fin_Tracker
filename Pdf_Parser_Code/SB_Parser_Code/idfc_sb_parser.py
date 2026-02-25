import os
import re
from glob import glob
from datetime import datetime

import pandas as pd
import pdfplumber
from openpyxl import load_workbook

PROJECT_DIR = os.path.expanduser(
    "~/Library/CloudStorage/OneDrive-Personal/Personal/Finance/projects/Monthly_Fin_Tracker"
)
BASE_DIR = os.path.join(PROJECT_DIR, "Bank_Statements", "SB_Statements")
OUTPUT_FILE = os.path.join(PROJECT_DIR, "Output", "idfc_summary.xlsx")


def clean_text(value):
    if value is None or pd.isna(value):
        return ""
    return re.sub(r"\s+", " ", str(value).strip())


def parse_amount(value):
    if value is None:
        return None
    s = clean_text(value).replace(",", "")
    s = re.sub(r"\bCR\b|\bDR\b", "", s, flags=re.I).strip()
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def parse_date(value):
    s = clean_text(value)
    for fmt in ("%d-%m-%Y", "%d/%m/%Y", "%d %b %y", "%d %b %Y", "%d-%b-%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def format_period(d):
    return d.strftime("%b-%y") if d else "Unknown"


def normalize_match_text(value):
    return re.sub(r"[^a-z0-9]", "", clean_text(value).lower())


def tokenize(value):
    return [t for t in re.split(r"[^a-z0-9]+", clean_text(value).lower()) if t]


def unordered_token_match(keyword, description):
    kw_tokens = tokenize(keyword)
    desc_tokens = tokenize(description)
    if not kw_tokens:
        return False
    for kt in kw_tokens:
        if not any(dt == kt or dt.startswith(kt) for dt in desc_tokens):
            return False
    return True


def detect_pdf_type(pdf_path):
    text_pages = 0
    table_pages = 0
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            if clean_text(page.extract_text()):
                text_pages += 1
            if page.extract_tables() or []:
                table_pages += 1
    if text_pages and table_pages:
        return "hybrid(text+table)"
    if text_pages:
        return "text"
    return "image/ocr-needed"


def resolve_mapping_file():
    ref_dir = os.path.join(PROJECT_DIR, "Reference Documents")
    candidates = glob(os.path.join(ref_dir, "Merchant category mapping*.xlsx"))
    if not candidates:
        raise FileNotFoundError("No mapping file found matching 'Merchant category mapping*.xlsx'")
    candidates.sort(key=lambda p: os.path.getmtime(p), reverse=True)
    return candidates[0]


def is_idfc_pdf(pdf_path):
    name = os.path.basename(pdf_path).lower()
    if "idfc" in name:
        return True
    try:
        with pdfplumber.open(pdf_path) as pdf:
            t = "\n".join((page.extract_text() or "") for page in pdf.pages[:2]).upper()
        return "IDFC FIRST BANK" in t or "IDFC BANK" in t
    except Exception:
        return False


def detect_idfc_account_name(pdf_path):
    try:
        with pdfplumber.open(pdf_path) as pdf:
            txt = "\n".join((p.extract_text() or "") for p in pdf.pages[:2]).upper()
    except Exception:
        return "IDFC"
    if "PRIYANKA JAIN" in txt:
        return "PJ IDFC"
    if "ABHISHEK JAIN" in txt:
        return "AJ IDFC"
    return "IDFC"


def is_noise_line(line):
    u = line.upper()
    return any(
        x in u
        for x in (
            "CONSOLIDATED STATEMENT",
            "STATEMENT PERIOD",
            "SUMMARY OF YOUR RELATIONSHIP",
            "ACCOUNT TYPE BALANCE",
            "TRANSACTION ACCOUNTS",
            "CUSTOMER ID",
            "CUSTOMER NAME",
            "ACCOUNT NAME",
            "ACCOUNT BRANCH",
            "BRANCH ADDRESS",
            "NOMINATION REGISTERED",
            "IFSC ",
            "MICR ",
            "ACCOUNT STATUS",
            "CURRENCY INR",
            "RAIPUR",
            "TOWER",
            "COLLEGE",
            "ROAD",
            "ACCOUNT OPENING DATE",
            "REGISTERED OFFICE",
            "PAGE ",
            "IF YOUR AADHAAR",
            "TO UPDATE, PLEASE VISIT",
            "DATE AND TIME VALUE DATE",
            "REF/CHEQUE",
            "(INR)",
        )
    )


def is_txn_prefix_line(line):
    return re.match(
        r"^(IMPS|BILLPAY|ATM|UPI|NEFT|RTGS|POS|ECOM|ACH|NACH|MONTHLY|WITHDRAWAL|CASH)\b",
        line.upper(),
    ) is not None


def append_fragment(desc, frag):
    desc = clean_text(desc)
    frag = clean_text(frag)
    if not frag:
        return desc
    if (
        desc
        and " " not in frag
        and re.search(r"[A-Za-z]$", desc)
        and re.match(r"^[a-z]{3,}[A-Za-z0-9/-]*$", frag)
    ):
        return f"{desc}{frag}"
    return f"{desc} {frag}".strip()


def parse_idfc_transactions(pdf_path):
    records = []
    prev_balance = None
    carry_prefix = []
    current_txn = None

    def flush_current():
        nonlocal current_txn, prev_balance
        if not current_txn:
            return
        line = current_txn["line"]
        nums = re.findall(r"\d{1,3}(?:,\d{2,3})*\.\d{2}", line)
        dates = re.findall(r"\b\d{2}\s+[A-Za-z]{3}\s+\d{2}\b", line)
        if len(nums) < 2 or not dates:
            current_txn = None
            return

        date = parse_date(dates[-1]) or parse_date(dates[0])
        if not date:
            current_txn = None
            return

        amount_abs = parse_amount(nums[-2])
        balance = parse_amount(nums[-1])
        if amount_abs is None or balance is None:
            current_txn = None
            return

        desc_tail = re.sub(
            r"^\d{2}\s+[A-Za-z]{3}\s+\d{2}(?:\s+\d{2}:\d{2})?(?:\s+\d{2}\s+[A-Za-z]{3}\s+\d{2})?\s+",
            "",
            line,
        )
        desc_tail = re.sub(
            r"\s+" + re.escape(nums[-2]) + r"\s+" + re.escape(nums[-1]) + r"\s*(?:CR|DR)?\s*$",
            "",
            desc_tail,
            flags=re.I,
        ).strip()
        if re.fullmatch(r"(?:\d{1,3}(?:,\d{2,3})*\.\d{2}\s*)+(?:CR|DR)?", desc_tail, flags=re.I):
            desc_tail = ""

        desc = ""
        for part in current_txn["pre"]:
            desc = append_fragment(desc, part)
        desc = append_fragment(desc, desc_tail)
        for part in current_txn["post"]:
            desc = append_fragment(desc, part)

        if prev_balance is None:
            prev_balance = balance
            current_txn = None
            return

        amount = amount_abs if balance > prev_balance else -amount_abs
        prev_balance = balance

        records.append(
            {
                "Period": format_period(date),
                "Date": date,
                "Account": "IDFC",
                "Description": clean_text(desc),
                "Amount": amount,
                "Balance": balance,
            }
        )
        current_txn = None

    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""
            lines = [clean_text(l) for l in text.split("\n") if clean_text(l)]
            for line in lines:
                if re.search(r"\bopening balance\b", line, re.I):
                    nums = re.findall(r"\d{1,3}(?:,\d{2,3})*\.\d{2}", line)
                    if nums:
                        prev_balance = parse_amount(nums[-1])
                    continue

                if is_noise_line(line):
                    continue

                if re.match(r"^\d{2}\s+[A-Za-z]{3}\s+\d{2}\s+\d{2}:\d{2}\s+\d{2}\s+[A-Za-z]{3}\s+\d{2}\b", line):
                    flush_current()
                    current_txn = {"pre": carry_prefix, "line": line, "post": []}
                    carry_prefix = []
                    continue

                if current_txn is not None:
                    if is_txn_prefix_line(line):
                        carry_prefix.append(line)
                    else:
                        current_txn["post"].append(line)
                elif is_txn_prefix_line(line):
                    carry_prefix.append(line)

    flush_current()
    return records


def load_idfc_mapping_rules():
    mapping_file = resolve_mapping_file()
    wb = load_workbook(mapping_file, data_only=True)
    sheet_name = "SB Mapping" if "SB Mapping" in wb.sheetnames else "SB Merchant category mapping"
    ws = wb[sheet_name]

    map_i_col = clean_text(ws.cell(row=1, column=9).value) or "Map Col I"
    map_j_col = clean_text(ws.cell(row=1, column=10).value) or "Map Col J"

    idfc_rules = []
    default_rules = []
    default_fallback = ("", "", "Uncategorized", "Uncategorized", "Uncategorized", "Unknown")

    for r in range(2, ws.max_row + 1):
        bank = clean_text(ws.cell(row=r, column=1).value).lower()
        keyword = clean_text(ws.cell(row=r, column=2).value)
        mapped = (
            clean_text(ws.cell(row=r, column=9).value),
            clean_text(ws.cell(row=r, column=10).value),
            clean_text(ws.cell(row=r, column=3).value) or "Uncategorized",
            clean_text(ws.cell(row=r, column=4).value) or "Uncategorized",
            clean_text(ws.cell(row=r, column=5).value) or "Uncategorized",
            clean_text(ws.cell(row=r, column=6).value) or "Unknown",
        )

        if bank == "idfc" and keyword:
            idfc_rules.append((keyword.lower(),) + mapped)
        elif bank == "default" and keyword:
            default_rules.append((keyword.lower(),) + mapped)
        elif bank == "default" and not keyword:
            default_fallback = mapped

    return idfc_rules, default_rules, default_fallback, map_i_col, map_j_col


def resolve_directional_mapping(matches, amount):
    if amount is None or not matches:
        return matches[0]
    is_withdrawal = amount < 0
    for item in matches:
        _, map_i_val, map_j_val, mode, exp_type, merch_cat, store_name = item
        txt = f"{map_i_val} {map_j_val} {mode} {exp_type} {merch_cat} {store_name}".lower()
        if is_withdrawal and re.search(r"\bsbi\b.*\bto\b.*\bidfc\b", txt):
            return item
        if (not is_withdrawal) and re.search(r"\bidfc\b.*\bto\b.*\bsbi\b", txt):
            return item
    return matches[0]


def _match_from_rules(description, amount, rules):
    desc = clean_text(description).lower()
    desc_norm = normalize_match_text(desc)
    matches = []
    for keyword, map_i_val, map_j_val, mode, exp_type, merch_cat, store_name in rules:
        key_norm = normalize_match_text(keyword)
        if not key_norm:
            continue
        if keyword in desc or key_norm in desc_norm or unordered_token_match(keyword, desc):
            matches.append((keyword, map_i_val, map_j_val, mode, exp_type, merch_cat, store_name))
    if matches:
        chosen = resolve_directional_mapping(matches, amount)
        _, map_i_val, map_j_val, mode, exp_type, merch_cat, store_name = chosen
        return map_i_val, map_j_val, mode, exp_type, merch_cat, store_name
    return None


def classify_idfc_row(description, amount, idfc_rules, default_rules, default_fallback):
    m = _match_from_rules(description, amount, idfc_rules)
    if m is not None:
        return m
    m = _match_from_rules(description, amount, default_rules)
    if m is not None:
        return m
    return default_fallback


def format_sheet(workbook, worksheet, df):
    nrows, ncols = df.shape
    if ncols == 0:
        return

    header_fmt = workbook.add_format({"bold": True, "bg_color": "#F4B183", "border": 1})
    cell_fmt = workbook.add_format({"border": 1})
    amt_fmt = workbook.add_format({"border": 1, "num_format": "#,##0.00"})

    for c, col in enumerate(df.columns):
        worksheet.write(0, c, col, header_fmt)

    numeric_cols = {"Amount", "Balance"}
    for r in range(1, nrows + 1):
        for c, col in enumerate(df.columns):
            val = df.iloc[r - 1, c]
            if pd.isna(val):
                worksheet.write_blank(r, c, None, cell_fmt)
            elif col in numeric_cols:
                try:
                    worksheet.write_number(r, c, float(val), amt_fmt)
                except Exception:
                    worksheet.write(r, c, val, cell_fmt)
            else:
                worksheet.write(r, c, val, cell_fmt)

    worksheet.freeze_panes(1, 0)
    worksheet.autofilter(0, 0, nrows, ncols - 1)
    for c, col in enumerate(df.columns):
        max_len = len(str(col))
        if nrows:
            max_len = max(max_len, int(df[col].astype(str).str.len().quantile(0.95)))
        worksheet.set_column(c, c, min(max_len + 2, 60))


def main():
    print("=" * 70)
    print("IDFC SB PARSER")
    print("=" * 70)

    pdf_paths = []
    for root, _, files in os.walk(BASE_DIR):
        for f in files:
            if f.lower().endswith(".pdf"):
                pdf_paths.append(os.path.join(root, f))

    idfc_pdfs = [p for p in sorted(pdf_paths) if is_idfc_pdf(p)]
    if not idfc_pdfs:
        print("No IDFC PDF found.")
        return

    all_records = []
    for pdf_path in idfc_pdfs:
        ptype = detect_pdf_type(pdf_path)
        print(f"\nProcessing: {os.path.basename(pdf_path)}")
        print(f"Detected PDF type: {ptype}")
        recs = parse_idfc_transactions(pdf_path)
        account_name = detect_idfc_account_name(pdf_path)
        for r in recs:
            r["Account"] = account_name
        print(f"Extracted transactions: {len(recs)}")
        all_records.extend(recs)

    if not all_records:
        print("No transactions extracted.")
        return

    df = pd.DataFrame(all_records)
    df["Date"] = pd.to_datetime(df["Date"], errors="coerce").dt.strftime("%d-%b-%Y")
    df = df[["Period", "Date", "Account", "Description", "Amount", "Balance"]]

    idfc_rules, default_rules, default_fallback, _, _ = load_idfc_mapping_rules()
    mapped = df.apply(
        lambda r: pd.Series(classify_idfc_row(r["Description"], r["Amount"], idfc_rules, default_rules, default_fallback)),
        axis=1,
    )
    mapped.columns = ["_Map Col I", "_Map Col J", "Mode", "Expense Type", "Merchant Category", "Store Name"]

    summary_df = pd.concat(
        [
            df[["Period", "Date", "Account", "Description", "Amount", "Balance"]],
            mapped[["Mode", "Expense Type", "Merchant Category", "Store Name"]],
        ],
        axis=1,
    )

    with pd.ExcelWriter(OUTPUT_FILE, engine="xlsxwriter") as writer:
        df.to_excel(writer, sheet_name="IDFC Transactions", index=False)
        summary_df.to_excel(writer, sheet_name="IDFC Categorized Summary", index=False)
        wb = writer.book
        format_sheet(wb, writer.sheets["IDFC Transactions"], df)
        format_sheet(wb, writer.sheets["IDFC Categorized Summary"], summary_df)

    print("\n" + "=" * 70)
    print("Completed")
    print(f"IDFC PDFs: {len(idfc_pdfs)}")
    print(f"Transactions: {len(df)}")
    print(f"Output: {OUTPUT_FILE}")
    print("=" * 70)


if __name__ == "__main__":
    main()
