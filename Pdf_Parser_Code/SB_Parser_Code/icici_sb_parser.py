import os
import re
from glob import glob
from datetime import datetime

import pandas as pd
import pdfplumber
from openpyxl import load_workbook

PROJECT_DIR = os.path.expanduser(
    "~/Library/CloudStorage/OneDrive-Personal/Personal/Finance/projects/Monthly_Fin_Tracker"
)
BASE_DIR = os.path.join(PROJECT_DIR, "Bank_Statements", "SB_Statements")
OUTPUT_FILE = os.path.join(PROJECT_DIR, "Output", "icici_summary.xlsx")


def clean_text(value):
    if value is None or pd.isna(value):
        return ""
    return re.sub(r"\s+", " ", str(value).strip())


def parse_amount(value):
    if value is None:
        return None
    s = clean_text(value).replace(",", "")
    s = re.sub(r"\bCR\b|\bDR\b", "", s, flags=re.I).strip()
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def parse_date(value):
    s = clean_text(value)
    for fmt in ("%d-%m-%Y", "%d/%m/%Y", "%d %b %y", "%d %b %Y", "%d-%b-%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def format_period(d):
    return d.strftime("%b-%y") if d else "Unknown"


def normalize_match_text(value):
    return re.sub(r"[^a-z0-9]", "", clean_text(value).lower())


def tokenize(value):
    return [t for t in re.split(r"[^a-z0-9]+", clean_text(value).lower()) if t]


def unordered_token_match(keyword, description):
    kw_tokens = tokenize(keyword)
    desc_tokens = tokenize(description)
    if not kw_tokens:
        return False
    for kt in kw_tokens:
        if not any(dt == kt or dt.startswith(kt) for dt in desc_tokens):
            return False
    return True


def detect_pdf_type(pdf_path):
    text_pages = 0
    table_pages = 0
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            if clean_text(page.extract_text()):
                text_pages += 1
            if page.extract_tables() or []:
                table_pages += 1
    if text_pages and table_pages:
        return "hybrid(text+table)"
    if text_pages:
        return "text"
    return "image/ocr-needed"


def resolve_mapping_file():
    ref_dir = os.path.join(PROJECT_DIR, "Reference Documents")
    candidates = glob(os.path.join(ref_dir, "Merchant category mapping*.xlsx"))
    if not candidates:
        raise FileNotFoundError("No mapping file found matching 'Merchant category mapping*.xlsx'")
    candidates.sort(key=lambda p: os.path.getmtime(p), reverse=True)
    return candidates[0]


def is_icici_pdf(pdf_path):
    name = os.path.basename(pdf_path).lower()
    if "icici" in name:
        return True
    try:
        with pdfplumber.open(pdf_path) as pdf:
            t = "\n".join((page.extract_text() or "") for page in pdf.pages[:2]).upper()
        return "ICICI BANK" in t
    except Exception:
        return False


def is_noise_line(line):
    u = line.upper()
    return any(
        x in u
        for x in (
            "ACCOUNT TYPE ACCOUNT BALANCE",
            "DATE MODE PARTICULARS",
            "DATE OF TAX WITHHELD",
            "OPENING BALANCE (CUMULATIVE)",
            "ACCOUNT NUMBER",
            "NOMINATION",
            "CREDITED (INR)",
            "TAX WITHHELD",
            "PAGE ",
        )
    )


def is_continuation_noise(line):
    u = line.upper()
    if "TOTAL:" in u:
        return True
    if "STATEMENT OF TRANSACTIONS" in u:
        return True
    if "SUMMARY OF TDS/INTEREST" in u:
        return True
    if "CLOSING BALANCE (CUMULATIVE)" in u:
        return True
    if "DATE OF TAX WITHHELD" in u:
        return True
    if "DATE MODE PARTICULARS" in u:
        return True
    return False


def is_mode_only_text(text):
    t = clean_text(text).upper()
    if not t:
        return False
    known_modes = {
        "MOBILE BANKING",
        "INTERNET BANKING",
        "BRANCH BANKING",
        "ATM",
        "POS",
        "UPI",
        "RTGS",
        "NEFT",
        "IMPS",
    }
    if t in known_modes:
        return True
    # Generic mode-ish token: mostly letters/spaces and no separators/digits.
    return re.fullmatch(r"[A-Z ]{3,30}", t) is not None


def is_next_txn_prefix_line(line):
    u = clean_text(line).upper()
    return (
        u.startswith("MMT/")
        or u.startswith("RTGS/")
        or u.startswith("BIL/")
        or u.startswith("UPI")
        or u.startswith("NEFT")
        or u.startswith("IMPS")
        or u.startswith("DCARDFEE")
        or re.match(r"^\d{12,}:", u) is not None
    )


def extract_icici_section_account(line):
    l = clean_text(line)
    if not re.search(r"statement of transactions", l, re.I):
        return None
    m = re.search(
        r"(?:savings\s+account|account(?:\s+no\.?)?)\s+X+\s*(\d{4})\b",
        l,
        re.I,
    )
    if not m:
        m = re.search(
            r"(?:savings\s+account|account(?:\s+no\.?)?)\s+X+(\d{4})\b",
            l,
            re.I,
        )
    if not m:
        m = re.search(
            r"(?:savings\s+account|account(?:\s+no\.?)?)\s+(\d{4,})\b",
            l,
            re.I,
        )
    if not m:
        return None
    last4 = m.group(1)[-4:]
    return f"XX{last4}"


def display_icici_account(account_code):
    account_map = {
        "XX0915": "ICICI_NRO",
        "XX5948": "ICICI_NRE",
    }
    return account_map.get(account_code, account_code)


def parse_icici_transactions(pdf_path):
    records = []
    last_account = "ICICI"
    last_balance_by_account = {}

    txn_line_re = re.compile(
        r"^(\d{2}-\d{2}-\d{4})\s+(.+?)\s+(\d{1,3}(?:,\d{2,3})*\.\d{2})\s+(\d{1,3}(?:,\d{2,3})*\.\d{2})$"
    )

    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""
            all_lines = [clean_text(l) for l in text.split("\n") if clean_text(l)]

            sections = []
            in_txn_section = False
            section_account = last_account
            curr_section = []
            for idx, line in enumerate(all_lines):
                detected_account = extract_icici_section_account(line)
                if detected_account:
                    last_account = detected_account
                    if not in_txn_section:
                        section_account = detected_account
                    continue

                if re.search(r"\bDATE\s+MODE\s+PARTICULARS\b", line, re.I):
                    if curr_section:
                        sections.append((section_account, curr_section))
                        curr_section = []
                    section_account = last_account
                    in_txn_section = True
                    continue

                if re.search(r"^Total:", line, re.I):
                    if curr_section:
                        sections.append((section_account, curr_section))
                        curr_section = []
                    in_txn_section = False
                    continue

                if re.search(r"\bDATE OF TAX WITHHELD\b", line, re.I) or re.search(
                    r"\bSummary of TDS/Interest\b",
                    line,
                    re.I,
                ):
                    in_txn_section = False
                    if curr_section:
                        sections.append((section_account, curr_section))
                        curr_section = []
                    continue

                if in_txn_section:
                    curr_section.append((line, idx))
            if curr_section:
                sections.append((section_account, curr_section))

            if not sections:
                continue

            for section_account, section_lines in sections:
                prev_balance_section = last_balance_by_account.get(section_account)
                # B/F lines define opening balance anchors for subsequent transaction lines.
                bf_anchors = []
                for i, (ln, _) in enumerate(section_lines):
                    if re.match(r"^\d{2}-\d{2}-\d{4}\s+B/F\b", ln, re.I):
                        nums = re.findall(r"\d{1,3}(?:,\d{2,3})*\.\d{2}", ln)
                        if nums:
                            bal = parse_amount(nums[-1])
                            if bal is not None:
                                bf_anchors.append((i, bal))

                date_idxs = [i for i, (ln, _) in enumerate(section_lines) if txn_line_re.match(ln)]
                consumed = set()
                bf_ptr = 0

                for pos, idx in enumerate(date_idxs):
                    while bf_ptr < len(bf_anchors) and bf_anchors[bf_ptr][0] < idx:
                        prev_balance_section = bf_anchors[bf_ptr][1]
                        bf_ptr += 1

                    line, line_idx = section_lines[idx]
                    line_account = section_account or "ICICI"
                    m = txn_line_re.match(line)
                    if not m:
                        continue

                    d = parse_date(m.group(1))
                    if not d:
                        continue
                    nums = re.findall(r"\d{1,3}(?:,\d{2,3})*\.\d{2}", line)
                    if len(nums) < 2:
                        continue
                    amount_abs = parse_amount(nums[-2])
                    balance = parse_amount(nums[-1])
                    if amount_abs is None or balance is None:
                        continue

                    prev_idx = date_idxs[pos - 1] if pos > 0 else -1
                    next_idx = date_idxs[pos + 1] if pos + 1 < len(date_idxs) else len(section_lines)

                    middle = re.sub(r"^\d{2}-\d{2}-\d{4}\s+", "", line)
                    middle = re.sub(
                        r"\s+" + re.escape(nums[-2]) + r"\s+" + re.escape(nums[-1]) + r"\s*$",
                        "",
                        middle,
                    ).strip()

                    # Prefix lines immediately above current dated line (not already consumed).
                    prefix_parts = []
                    j = idx - 1
                    while j > prev_idx:
                        cand, _ = section_lines[j]
                        if j in consumed or is_noise_line(cand) or is_continuation_noise(cand):
                            j -= 1
                            continue
                        if re.match(r"^\d{2}-\d{2}-\d{4}\b", cand):
                            break
                        # Avoid stealing short suffix from previous txn.
                        if re.fullmatch(r"(Bank(?: Ltd)?|Ltd)", cand, flags=re.I):
                            break
                        if is_next_txn_prefix_line(cand):
                            prefix_parts.append(cand)
                            consumed.add(j)
                            j -= 1
                            continue
                        break
                    prefix_parts.reverse()

                    # Suffix lines immediately below current dated line, but only safe continuations.
                    suffix_parts = []
                    j = idx + 1
                    while j < next_idx:
                        cand, _ = section_lines[j]
                        if j in consumed or is_noise_line(cand) or is_continuation_noise(cand):
                            j += 1
                            continue
                        if re.match(r"^\d{2}-\d{2}-\d{4}\b", cand):
                            break
                        if is_next_txn_prefix_line(cand):
                            break
                        if re.fullmatch(r"(Bank(?: Ltd)?|Ltd)", cand, flags=re.I):
                            suffix_parts.append(cand)
                            consumed.add(j)
                            j += 1
                            continue
                        # Generic continuation detail line for current txn (not mode, not next txn prefix).
                        if "/" in cand or ":" in cand or re.search(r"[A-Za-z]{3,}\d", cand):
                            suffix_parts.append(cand)
                            consumed.add(j)
                            j += 1
                            continue
                        break

                    desc_parts = []
                    desc_parts.extend(prefix_parts)
                    if not (prefix_parts and is_mode_only_text(middle)):
                        desc_parts.append(middle)
                    desc_parts.extend(suffix_parts)
                    desc = clean_text(" ".join([p for p in desc_parts if clean_text(p)]))

                    if prev_balance_section is None:
                        prev_balance_section = balance
                        last_balance_by_account[section_account] = balance
                        continue

                    amount = amount_abs if balance > prev_balance_section else -amount_abs
                    prev_balance_section = balance
                    last_balance_by_account[section_account] = balance
                    acct = display_icici_account(line_account or "ICICI")
                    records.append(
                        {
                            "Period": format_period(d),
                            "Date": d,
                            "Account": acct,
                            "Description": desc,
                            "Amount": amount,
                            "Balance": balance,
                        }
                    )
    return records


def load_icici_mapping_rules():
    mapping_file = resolve_mapping_file()
    wb = load_workbook(mapping_file, data_only=True)
    sheet_name = "SB Mapping" if "SB Mapping" in wb.sheetnames else "SB Merchant category mapping"
    ws = wb[sheet_name]

    map_i_col = clean_text(ws.cell(row=1, column=9).value) or "Map Col I"
    map_j_col = clean_text(ws.cell(row=1, column=10).value) or "Map Col J"

    icici_rules = []
    default_rules = []
    default_fallback = ("", "", "Uncategorized", "Uncategorized", "Uncategorized", "Unknown")

    for r in range(2, ws.max_row + 1):
        bank = clean_text(ws.cell(row=r, column=1).value).lower()
        keyword = clean_text(ws.cell(row=r, column=2).value)
        mapped = (
            clean_text(ws.cell(row=r, column=9).value),
            clean_text(ws.cell(row=r, column=10).value),
            clean_text(ws.cell(row=r, column=3).value) or "Uncategorized",
            clean_text(ws.cell(row=r, column=4).value) or "Uncategorized",
            clean_text(ws.cell(row=r, column=5).value) or "Uncategorized",
            clean_text(ws.cell(row=r, column=6).value) or "Unknown",
        )

        if bank == "icici" and keyword:
            icici_rules.append((keyword.lower(),) + mapped)
        elif bank == "default" and keyword:
            default_rules.append((keyword.lower(),) + mapped)
        elif bank == "default" and not keyword:
            default_fallback = mapped

    return icici_rules, default_rules, default_fallback, map_i_col, map_j_col


def _match_from_rules(description, rules):
    desc = clean_text(description).lower()
    desc_norm = normalize_match_text(desc)
    for keyword, map_i_val, map_j_val, mode, exp_type, merch_cat, store_name in rules:
        key_norm = normalize_match_text(keyword)
        if not key_norm:
            continue
        if keyword in desc or key_norm in desc_norm or unordered_token_match(keyword, desc):
            return map_i_val, map_j_val, mode, exp_type, merch_cat, store_name
    return None


def classify_icici_row(description, icici_rules, default_rules, default_fallback):
    m = _match_from_rules(description, icici_rules)
    if m is not None:
        return m
    m = _match_from_rules(description, default_rules)
    if m is not None:
        return m
    return default_fallback


def format_sheet(workbook, worksheet, df):
    nrows, ncols = df.shape
    if ncols == 0:
        return

    header_fmt = workbook.add_format({"bold": True, "bg_color": "#F4B183", "border": 1})
    cell_fmt = workbook.add_format({"border": 1})
    amt_fmt = workbook.add_format({"border": 1, "num_format": "#,##0.00"})

    for c, col in enumerate(df.columns):
        worksheet.write(0, c, col, header_fmt)

    numeric_cols = {"Amount", "Balance"}
    for r in range(1, nrows + 1):
        for c, col in enumerate(df.columns):
            val = df.iloc[r - 1, c]
            if pd.isna(val):
                worksheet.write_blank(r, c, None, cell_fmt)
            elif col in numeric_cols:
                try:
                    worksheet.write_number(r, c, float(val), amt_fmt)
                except Exception:
                    worksheet.write(r, c, val, cell_fmt)
            else:
                worksheet.write(r, c, val, cell_fmt)

    worksheet.freeze_panes(1, 0)
    worksheet.autofilter(0, 0, nrows, ncols - 1)
    for c, col in enumerate(df.columns):
        max_len = len(str(col))
        if nrows:
            max_len = max(max_len, int(df[col].astype(str).str.len().quantile(0.95)))
        worksheet.set_column(c, c, min(max_len + 2, 60))


def main():
    print("=" * 70)
    print("ICICI SB PARSER")
    print("=" * 70)

    pdf_paths = []
    for root, _, files in os.walk(BASE_DIR):
        for f in files:
            if f.lower().endswith(".pdf"):
                pdf_paths.append(os.path.join(root, f))

    icici_pdfs = [p for p in sorted(pdf_paths) if is_icici_pdf(p)]
    if not icici_pdfs:
        print("No ICICI PDF found.")
        return

    all_records = []
    for pdf_path in icici_pdfs:
        ptype = detect_pdf_type(pdf_path)
        print(f"\nProcessing: {os.path.basename(pdf_path)}")
        print(f"Detected PDF type: {ptype}")
        recs = parse_icici_transactions(pdf_path)
        print(f"Extracted transactions: {len(recs)}")
        all_records.extend(recs)

    if not all_records:
        print("No transactions extracted.")
        return

    df = pd.DataFrame(all_records)
    df["Date"] = pd.to_datetime(df["Date"], errors="coerce").dt.strftime("%d-%b-%Y")
    df = df[["Period", "Date", "Account", "Description", "Amount", "Balance"]]

    icici_rules, default_rules, default_fallback, _, _ = load_icici_mapping_rules()
    mapped = df["Description"].apply(
        lambda d: pd.Series(classify_icici_row(d, icici_rules, default_rules, default_fallback))
    )
    mapped.columns = ["_Map Col I", "_Map Col J", "Mode", "Expense Type", "Merchant Category", "Store Name"]

    summary_df = pd.concat(
        [
            df[["Period", "Date", "Account", "Description", "Amount", "Balance"]],
            mapped[["Mode", "Expense Type", "Merchant Category", "Store Name"]],
        ],
        axis=1,
    )

    with pd.ExcelWriter(OUTPUT_FILE, engine="xlsxwriter") as writer:
        df.to_excel(writer, sheet_name="ICICI Transactions", index=False)
        summary_df.to_excel(writer, sheet_name="ICICI Categorized Summary", index=False)
        wb = writer.book
        format_sheet(wb, writer.sheets["ICICI Transactions"], df)
        format_sheet(wb, writer.sheets["ICICI Categorized Summary"], summary_df)

    print("\n" + "=" * 70)
    print("Completed")
    print(f"ICICI PDFs: {len(icici_pdfs)}")
    print(f"Transactions: {len(df)}")
    print(f"Output: {OUTPUT_FILE}")
    print("=" * 70)


if __name__ == "__main__":
    main()
