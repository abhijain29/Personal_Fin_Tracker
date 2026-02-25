import os
import re
from glob import glob
from datetime import datetime

import pandas as pd
import pdfplumber
from openpyxl import load_workbook
from openpyxl.styles import Border, Font, PatternFill, Side

PROJECT_DIR = os.path.expanduser(
    "~/Library/CloudStorage/OneDrive-Personal/Personal/Finance/projects/Monthly_Fin_Tracker"
)
BASE_DIR = os.path.join(PROJECT_DIR, "Bank_Statements", "SB_Statements")
OUTPUT_FILE = os.path.join(PROJECT_DIR, "Output", "axis_summary.xlsx")
TEMPLATE_FILE = os.path.join(PROJECT_DIR, "Reference Documents", "axis_summary_template.xlsx")


def clean_text(value):
    if value is None or pd.isna(value):
        return ""
    return re.sub(r"\s+", " ", str(value).strip())


def parse_amount(value):
    if value is None:
        return None
    s = clean_text(value).replace(",", "")
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def parse_date(value):
    s = clean_text(value)
    for fmt in ("%d-%m-%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def format_period(d):
    return d.strftime("%b-%y") if d else "Unknown"


def normalize_match_text(value):
    return re.sub(r"[^a-z0-9]", "", clean_text(value).lower())


def tokenize(value):
    return [t for t in re.split(r"[^a-z0-9]+", clean_text(value).lower()) if t]


def unordered_token_match(keyword, description):
    kw_tokens = tokenize(keyword)
    desc_tokens = tokenize(description)
    if not kw_tokens:
        return False
    for kt in kw_tokens:
        if not any(dt == kt or dt.startswith(kt) for dt in desc_tokens):
            return False
    return True


def detect_pdf_type(pdf_path):
    text_pages = 0
    table_pages = 0
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            if clean_text(page.extract_text()):
                text_pages += 1
            if page.extract_tables() or []:
                table_pages += 1
    if text_pages and table_pages:
        return "hybrid(text+table)"
    if text_pages:
        return "text"
    return "image/ocr-needed"


def resolve_mapping_file():
    ref_dir = os.path.join(PROJECT_DIR, "Reference Documents")
    candidates = glob(os.path.join(ref_dir, "Merchant category mapping*.xlsx"))
    if not candidates:
        raise FileNotFoundError("No mapping file found matching 'Merchant category mapping*.xlsx'")
    candidates.sort(key=lambda p: os.path.getmtime(p), reverse=True)
    return candidates[0]


def is_axis_pdf(pdf_path):
    name = os.path.basename(pdf_path).lower()
    if "axis" in name:
        return True
    try:
        with pdfplumber.open(pdf_path) as pdf:
            t = "\n".join((page.extract_text() or "") for page in pdf.pages[:2]).upper()
        return "AXIS BANK" in t
    except Exception:
        return False


def is_non_txn_noise(line):
    u = line.upper()
    return any(
        x in u
        for x in (
            "STATEMENT OF AXIS",
            "CUST ID",
            "RELATIONSHIP MANAGER",
            "ACCOUNT NO.",
            "DATE TRANSACTION DETAILS",
            "WITHDRAWAL",
            "DEPOSITS",
            "LEGENDS USED IN THE STATEMENT",
            "DISCLAIMER",
            "REGISTERED OFFICE",
            "THIS IS A SYSTEM GENERATED",
            "BALANCE AS ON",
            "TOTAL ",
            "PAGE ",
        )
    )


def parse_axis_transactions(pdf_path):
    records = []
    in_section = False
    last_rec = None

    txn_line_re = re.compile(
        r"^(\d{2}-\d{2}-\d{4})\s+(.+?)\s+(\d{1,3}(?:,\d{2,3})*\.\d{2})\s+(\d{1,3}(?:,\d{2,3})*\.\d{2})\s+(\d{1,3}(?:,\d{2,3})*\.\d{2})$"
    )

    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""
            lines = [clean_text(l) for l in text.split("\n") if clean_text(l)]
            for line in lines:
                if re.search(r"\bOpening Balance\b", line, flags=re.I):
                    in_section = True
                    last_rec = None
                    continue

                if not in_section:
                    continue

                if re.search(r"\bClosing Balance\b", line, flags=re.I):
                    in_section = False
                    last_rec = None
                    continue

                m = txn_line_re.match(line)
                if m:
                    d = parse_date(m.group(1))
                    if not d:
                        continue

                    desc = clean_text(m.group(2))
                    wd = parse_amount(m.group(3)) or 0.0
                    dep = parse_amount(m.group(4)) or 0.0
                    bal = parse_amount(m.group(5))
                    amount = dep - wd

                    rec = {
                        "Period": format_period(d),
                        "Date": d,
                        "Account": "Axis",
                        "Description": desc,
                        "Amount": amount,
                        "Balance": bal,
                    }
                    records.append(rec)
                    last_rec = rec
                    continue

                # Continuation lines: append to previous description so no source text is lost.
                if last_rec is not None:
                    if re.match(r"^\d{2}-\d{2}-\d{4}$", line):
                        last_rec["Description"] = clean_text(f"{last_rec['Description']} {line}")
                        continue
                    if not is_non_txn_noise(line):
                        last_rec["Description"] = clean_text(f"{last_rec['Description']} {line}")

    return records


def load_axis_mapping_rules():
    mapping_file = resolve_mapping_file()
    wb = load_workbook(mapping_file, data_only=True)
    sheet_name = "SB Mapping" if "SB Mapping" in wb.sheetnames else "SB Merchant category mapping"
    ws = wb[sheet_name]

    map_i_col = clean_text(ws.cell(row=1, column=9).value) or "Map Col I"
    map_j_col = clean_text(ws.cell(row=1, column=10).value) or "Map Col J"

    axis_rules = []
    default_rules = []
    default_fallback = ("", "", "Uncategorized", "Uncategorized", "Uncategorized", "Unknown")

    for r in range(2, ws.max_row + 1):
        bank = clean_text(ws.cell(row=r, column=1).value).lower()
        keyword = clean_text(ws.cell(row=r, column=2).value)
        mapped = (
            clean_text(ws.cell(row=r, column=9).value),
            clean_text(ws.cell(row=r, column=10).value),
            clean_text(ws.cell(row=r, column=3).value) or "Uncategorized",
            clean_text(ws.cell(row=r, column=4).value) or "Uncategorized",
            clean_text(ws.cell(row=r, column=5).value) or "Uncategorized",
            clean_text(ws.cell(row=r, column=6).value) or "Unknown",
        )

        if bank == "axis" and keyword:
            axis_rules.append((keyword.lower(),) + mapped)
        elif bank == "default" and keyword:
            default_rules.append((keyword.lower(),) + mapped)
        elif bank == "default" and not keyword:
            default_fallback = mapped

    return axis_rules, default_rules, default_fallback, map_i_col, map_j_col


def _match_from_rules(description, rules):
    desc = clean_text(description).lower()
    desc_norm = normalize_match_text(desc)
    for keyword, map_i_val, map_j_val, mode, exp_type, merch_cat, store_name in rules:
        key_norm = normalize_match_text(keyword)
        if not key_norm:
            continue
        if (
            keyword in desc
            or key_norm in desc_norm
            or unordered_token_match(keyword, desc)
        ):
            return map_i_val, map_j_val, mode, exp_type, merch_cat, store_name
    return None


def classify_axis_description(description, axis_rules, default_rules, default_fallback):
    # 1) Axis bank-specific keyword rules
    m = _match_from_rules(description, axis_rules)
    if m is not None:
        return m
    # 2) Default keyword rules
    m = _match_from_rules(description, default_rules)
    if m is not None:
        return m
    # 3) Default fallback row
    return default_fallback


def clear_range(ws, max_col, from_row=1):
    max_row = ws.max_row
    if max_row < from_row:
        return
    for r in range(from_row, max_row + 1):
        for c in range(1, max_col + 1):
            ws.cell(r, c).value = None


def write_df_to_sheet(ws, df, max_col):
    clear_range(ws, max_col=max_col, from_row=1)

    header_fill = PatternFill(fill_type="solid", fgColor="F4B183")
    header_font = Font(bold=True)
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    amt_fmt = "#,##0.00"

    for c, col in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=c, value=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border

    numeric_cols = {"Amount", "Balance"}
    for r in range(2, len(df) + 2):
        for c, col in enumerate(df.columns, start=1):
            val = df.iloc[r - 2, c - 1]
            cell = ws.cell(row=r, column=c)
            if pd.isna(val):
                cell.value = None
            elif col in numeric_cols:
                try:
                    cell.value = float(val)
                    cell.number_format = amt_fmt
                except Exception:
                    cell.value = val
            else:
                cell.value = val
            cell.border = border

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{chr(64 + len(df.columns))}{max(1, len(df) + 1)}"

    for c, col in enumerate(df.columns, start=1):
        max_len = len(str(col))
        if len(df):
            max_len = max(max_len, int(df[col].astype(str).str.len().quantile(0.95)))
        ws.column_dimensions[chr(64 + c)].width = min(max_len + 2, 60)


def write_output_from_template(df, summary_df):
    if not os.path.exists(TEMPLATE_FILE):
        return False

    wb = load_workbook(TEMPLATE_FILE)
    ws_txn = wb["Axis Transactions"] if "Axis Transactions" in wb.sheetnames else wb.create_sheet("Axis Transactions")
    ws_summary = (
        wb["Axis Categorized Summary"]
        if "Axis Categorized Summary" in wb.sheetnames
        else wb.create_sheet("Axis Categorized Summary")
    )

    write_df_to_sheet(ws_txn, df, max_col=6)
    write_df_to_sheet(ws_summary, summary_df, max_col=10)

    wb.save(OUTPUT_FILE)
    return True



def main():
    print("=" * 70)
    print("AXIS SB PARSER")
    print("=" * 70)

    pdf_paths = []
    for root, _, files in os.walk(BASE_DIR):
        for f in files:
            if f.lower().endswith(".pdf"):
                pdf_paths.append(os.path.join(root, f))

    axis_pdfs = [p for p in sorted(pdf_paths) if is_axis_pdf(p)]
    if not axis_pdfs:
        print("No Axis PDF found.")
        return

    all_records = []
    for pdf_path in axis_pdfs:
        pdf_type = detect_pdf_type(pdf_path)
        print(f"\\nProcessing: {os.path.basename(pdf_path)}")
        print(f"Detected PDF type: {pdf_type}")
        recs = parse_axis_transactions(pdf_path)
        print(f"Extracted transactions: {len(recs)}")
        all_records.extend(recs)

    if not all_records:
        print("No transactions extracted.")
        return

    # Keep exact extraction order as it appears in PDF pages/lines.
    df = pd.DataFrame(all_records)
    df["Date"] = pd.to_datetime(df["Date"], errors="coerce").dt.strftime("%d-%b-%Y")
    df = df[["Period", "Date", "Account", "Description", "Amount", "Balance"]]

    axis_rules, default_rules, default_fallback, _, _ = load_axis_mapping_rules()
    mapped = df["Description"].apply(
        lambda d: pd.Series(classify_axis_description(d, axis_rules, default_rules, default_fallback))
    )
    mapped.columns = ["_Map Col I", "_Map Col J", "Mode", "Expense Type", "Merchant Category", "Store Name"]

    summary_df = pd.concat(
        [
            df[["Period", "Date", "Account", "Description"]],
            mapped[["Mode", "Expense Type", "Merchant Category", "Store Name"]],
            df[["Amount", "Balance"]],
        ],
        axis=1,
    )

    used_template = write_output_from_template(df, summary_df)
    if not used_template:
        with pd.ExcelWriter(OUTPUT_FILE, engine="xlsxwriter") as writer:
            df.to_excel(writer, sheet_name="Axis Transactions", index=False)
            summary_df.to_excel(writer, sheet_name="Axis Categorized Summary", index=False)

    print("\\n" + "=" * 70)
    print("Completed")
    print(f"Axis PDFs: {len(axis_pdfs)}")
    print(f"Transactions: {len(df)}")
    print(f"Output: {OUTPUT_FILE}")
    print("=" * 70)


if __name__ == "__main__":
    main()
